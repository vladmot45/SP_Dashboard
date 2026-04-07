"""
SP Trading Dashboard
All fixes applied:
 - Transport blank → Truck (everywhere, not just pie chart)
 - Price FCA from col 1 (exact match, same as reference code)
 - Currency-separated value totals (EUR contracts → EUR total only, etc.)
 - All filters (season/product/trader/year/month) applied to EVERY section including Wagi
 - Wagi unloaded matched by contract number (col 9 "№ контракта") + season filter
 - Value of unloaded quantity shown
 - Contract drill-down: click metric → see raw rows
"""

import re
import io
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import tempfile

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SP Trading Operations",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════════════════════════════════════════
# TRANSLATIONS
# ══════════════════════════════════════════════════════════════════════════════
T = {
    "EN": {
        "title": "SP Dashboard", "upload": "Upload Excel file", "save": "Save your changes",
        "download": "⬇  Download updated file", "updated": "Updated",
        "filters": "FILTERS", "season": "Season", "product": "Product",
        "trader": "Trader", "show_sections": "Show sections", "year": "Year", "month": "Month",
        "exchange_rates": "LIVE EXCHANGE RATES · NBP", "rate_date": "Rate date",
        "silo_occ": "SILO OCCUPANCY · CURRENT STOCK", "total": "TOTAL", "free": "Free",
        "contract_summary": "CONTRACT SUMMARY", "contracts": "Contracts",
        "sold_mt": "Sold (MT)", "issued_mt": "Issued (MT)", "left_mt": "Left to Issue (MT)",
        "avg_price": "Avg Price FCA", "value_sold": "VALUE OF SOLD CONTRACTS",
        "sold_eur": "Sold in EUR contracts", "sold_usd": "Sold in USD contracts",
        "sold_pln": "Sold in PLN contracts",
        "pln_help": "Sum of contracts where Currency = PLN (Price FCA × MT)",
        "transport_chart": "TRANSPORT MODE", "country_chart": "DESTINATION COUNTRY",
        "unload_schedule": "UNLOADING SCHEDULE",
        "schedule_desc": "Monthly planned deliveries — how much of each product should be unloaded and when.",
        "upcoming": "UPCOMING — NEXT 3 MONTHS", "no_upcoming": "No upcoming scheduled deliveries found.",
        "wagi_section": "UNLOADED · WAGI TOTAL",
        "wagi_desc": "Actual goods unloaded — matched by contract number from Wagi Total.",
        "total_unloaded": "Total Unloaded (MT)", "shipments": "Shipments",
        "unloaded_value": "Value of Unloaded",
        "by_product": "BY PRODUCT", "by_month": "BY MONTH", "recent_shipments": "RECENT SHIPMENTS",
        "no_data": "No data found for the selected filters.",
        "no_silo": "No silo data found for the selected products.",
        "upload_prompt": "👈  Upload your Excel file in the sidebar to get started.",
        "drill_title": "CONTRACT DETAILS",
        "drill_close": "Close details",
        "sections": ["Exchange Rates", "Silo Occupancy", "Contract Summary",
                     "Price Analysis", "Unloading Schedule", "Unloaded (Wagi Total)",
                     "Broker Report", "Buyer Summary"],
        "buyer": "Buyer", "broker": "Broker",
        "broker_report": "BROKER REPORT",
        "buyer_summary": "BUYER SUMMARY", "export_buyer": "📥 Export to Excel",
        "dl_buyer": "⬇ Download Buyer Report",
        "commission_rate": "Rate (€/MT)", "commission_total": "Total Commission (€)",
        "amount_due": "Amount Due (€)", "export_broker": "📥 Export to Excel",
        "dl_broker": "⬇ Download Report",
        "month_names": ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],
        "price_analysis": "PRICE ANALYSIS · FCA",
        "avg_price_eur": "Avg Price (EUR)",
        "avg_price_usd": "Avg Price (USD)",
        "price_chart_title": "Average FCA Price Over Time",
        "view_1y": "1 Year", "view_5y": "5 Years", "view_all": "All",
        "currency_toggle": "Display currency",
        "language": "Language",
    },
    "PL": {
        "title": "Panel główny", "upload": "Wgraj plik Excel", "save": "Zapisz zmiany",
        "download": "⬇  Pobierz zaktualizowany plik", "updated": "Zaktualizowano",
        "filters": "FILTRY", "season": "Sezon", "product": "Produkt",
        "trader": "Trader", "show_sections": "Pokaż sekcje", "year": "Rok", "month": "Miesiąc",
        "exchange_rates": "KURSY WALUT · NBP", "rate_date": "Data kursu",
        "silo_occ": "ZAJĘTOŚĆ SILOSÓW · AKTUALNY STAN", "total": "RAZEM", "free": "Wolne",
        "contract_summary": "PODSUMOWANIE KONTRAKTÓW", "contracts": "Kontrakty",
        "sold_mt": "Sprzedane (MT)", "issued_mt": "Wydane (MT)", "left_mt": "Pozostało do wydania (MT)",
        "avg_price": "Śr. cena FCA", "value_sold": "WARTOŚĆ SPRZEDANYCH KONTRAKTÓW",
        "sold_eur": "Sprzedane w EUR", "sold_usd": "Sprzedane w USD", "sold_pln": "Sprzedane w PLN",
        "pln_help": "Suma kontraktów gdzie Waluta = PLN (Cena FCA × MT)",
        "transport_chart": "ŚRODKI TRANSPORTU", "country_chart": "KRAJ DOCELOWY",
        "unload_schedule": "HARMONOGRAM ROZŁADUNKU",
        "schedule_desc": "Miesięczny plan dostaw.",
        "upcoming": "NADCHODZĄCE — NASTĘPNE 3 MIESIĄCE", "no_upcoming": "Brak nadchodzących dostaw.",
        "wagi_section": "ROZŁADOWANE · WAGI TOTAL",
        "wagi_desc": "Faktycznie rozładowane towary dopasowane przez numer kontraktu.",
        "total_unloaded": "Łącznie rozładowane (MT)", "shipments": "Dostawy",
        "unloaded_value": "Wartość rozładowanego",
        "by_product": "WEDŁUG PRODUKTU", "by_month": "WEDŁUG MIESIĄCA", "recent_shipments": "OSTATNIE DOSTAWY",
        "no_data": "Brak danych dla wybranych filtrów.", "no_silo": "Brak danych silosów.",
        "upload_prompt": "👈  Wgraj plik Excel w panelu bocznym.",
        "drill_title": "SZCZEGÓŁY KONTRAKTU", "drill_close": "Zamknij szczegóły",
        "sections": ["Kursy walut", "Zajętość silosów", "Podsumowanie kontraktów",
                     "Analiza cen", "Harmonogram rozładunku", "Rozładowane (Wagi Total)",
                     "Raport brokera", "Podsumowanie kupujących"],
        "buyer": "Kupujący", "broker": "Broker",
        "broker_report": "RAPORT BROKERA",
        "buyer_summary": "PODSUMOWANIE KUPUJĄCYCH", "export_buyer": "📥 Eksport do Excel",
        "dl_buyer": "⬇ Pobierz raport kupujących",
        "commission_rate": "Stawka (€/MT)", "commission_total": "Łączna prowizja (€)",
        "amount_due": "Należna kwota (€)", "export_broker": "📥 Eksport do Excel",
        "dl_broker": "⬇ Pobierz raport",
        "month_names": ["Sty","Lut","Mar","Kwi","Maj","Cze","Lip","Sie","Wrz","Paź","Lis","Gru"],
        "price_analysis": "ANALIZA CEN · FCA",
        "avg_price_eur": "Śr. cena (EUR)",
        "avg_price_usd": "Śr. cena (USD)",
        "price_chart_title": "Średnia cena FCA w czasie",
        "view_1y": "1 rok", "view_5y": "5 lat", "view_all": "Wszystko",
        "currency_toggle": "Waluta wyświetlania",
        "language": "Język",
    },
    "RU": {
        "title": "Дашборд", "upload": "Загрузить файл Excel", "save": "Сохранить изменения",
        "download": "⬇  Скачать обновлённый файл", "updated": "Обновлено",
        "filters": "ФИЛЬТРЫ", "season": "Сезон", "product": "Продукт",
        "trader": "Трейдер", "show_sections": "Показать разделы", "year": "Год", "month": "Месяц",
        "exchange_rates": "КУРСЫ ВАЛЮТ · НБП", "rate_date": "Дата курса",
        "silo_occ": "ЗАПОЛНЕННОСТЬ СИЛОСОВ", "total": "ИТОГО", "free": "Свободно",
        "contract_summary": "СВОДКА КОНТРАКТОВ", "contracts": "Контракты",
        "sold_mt": "Продано (МТ)", "issued_mt": "Отгружено (МТ)", "left_mt": "Осталось отгрузить (МТ)",
        "avg_price": "Ср. цена FCA", "value_sold": "СТОИМОСТЬ ПРОДАННЫХ КОНТРАКТОВ",
        "sold_eur": "Продано в EUR", "sold_usd": "Продано в USD", "sold_pln": "Продано в PLN",
        "pln_help": "Сумма контрактов где Валюта = PLN (Цена FCA × МТ)",
        "transport_chart": "ВИД ТРАНСПОРТА", "country_chart": "СТРАНА НАЗНАЧЕНИЯ",
        "unload_schedule": "ГРАФИК РАЗГРУЗКИ",
        "schedule_desc": "Ежемесячный план поставок.",
        "upcoming": "БЛИЖАЙШИЕ — СЛЕДУЮЩИЕ 3 МЕСЯЦА", "no_upcoming": "Нет предстоящих поставок.",
        "wagi_section": "РАЗГРУЖЕНО · WAGI TOTAL",
        "wagi_desc": "Фактически разгруженные товары, сопоставленные по номеру контракта.",
        "total_unloaded": "Всего разгружено (МТ)", "shipments": "Поставки",
        "unloaded_value": "Стоимость разгруженного",
        "by_product": "ПО ПРОДУКТУ", "by_month": "ПО МЕСЯЦАМ", "recent_shipments": "ПОСЛЕДНИЕ ПОСТАВКИ",
        "no_data": "Данные не найдены.", "no_silo": "Данные силосов не найдены.",
        "upload_prompt": "👈  Загрузите файл Excel в боковой панели.",
        "drill_title": "ДЕТАЛИ КОНТРАКТОВ", "drill_close": "Закрыть детали",
        "sections": ["Курсы валют", "Заполненность силосов", "Сводка контрактов",
                     "Анализ цен", "График разгрузки", "Разгружено (Wagi Total)",
                     "Отчёт брокера", "Сводка покупателей"],
        "buyer": "Покупатель", "broker": "Брокер",
        "broker_report": "ОТЧЁТ БРОКЕРА",
        "buyer_summary": "СВОДКА ПОКУПАТЕЛЕЙ", "export_buyer": "📥 Экспорт в Excel",
        "dl_buyer": "⬇ Скачать отчёт покупателей",
        "commission_rate": "Ставка (€/MT)", "commission_total": "Итого комиссия (€)",
        "amount_due": "К оплате (€)", "export_broker": "📥 Экспорт в Excel",
        "dl_broker": "⬇ Скачать отчёт",
        "month_names": ["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"],
        "price_analysis": "АНАЛИЗ ЦЕН · FCA",
        "avg_price_eur": "Ср. цена (EUR)",
        "avg_price_usd": "Ср. цена (USD)",
        "price_chart_title": "Средняя цена FCA во времени",
        "view_1y": "1 год", "view_5y": "5 лет", "view_all": "Все",
        "currency_toggle": "Валюта отображения",
        "language": "Язык",
    },
}

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600;700&family=IBM+Plex+Sans:wght@400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; font-size: 17px; color: #1a1a1a; }
    .stApp { background-color: #f5f4f0; }
    section[data-testid="stSidebar"] { background-color: #eceae4; border-right: 1px solid #c8c4bc; }
    section[data-testid="stSidebar"] p { font-size: 0.95rem !important; color: #505050 !important; }
    label, .stSelectbox label, .stMultiSelect label { font-size: 1.05rem !important; color: #2a2a2a !important; font-weight: 600 !important; }
    [data-testid="metric-container"] { background: #eceae4; border: 1px solid #c8c4bc; border-radius: 6px; padding: 20px 24px; cursor: pointer; }
    [data-testid="metric-container"] label { font-family: 'IBM Plex Mono', monospace !important; font-size: 0.85rem !important; color: #666666 !important; text-transform: uppercase; letter-spacing: 0.1em; font-weight: 700 !important; }
    [data-testid="metric-container"] [data-testid="stMetricValue"] { font-family: 'IBM Plex Mono', monospace; font-size: 2rem !important; color: #1a1a1a !important; font-weight: 700 !important; }
    .sec-hdr { font-family: 'IBM Plex Mono', monospace; font-size: 0.82rem; color: #1a5fa0; letter-spacing: 0.2em; text-transform: uppercase; border-bottom: 1px solid #c8c4bc; padding-bottom: 10px; margin-bottom: 22px; margin-top: 16px; font-weight: 700; }
    .sub-hdr { font-family: 'IBM Plex Mono', monospace; font-size: 0.95rem; color: #1a5fa0; letter-spacing: 0.1em; text-transform: uppercase; margin-bottom: 12px; font-weight: 700; }
    h1 { font-family: 'IBM Plex Mono', monospace; color: #1a1a1a; font-size: 2rem !important; font-weight: 700; letter-spacing: -0.02em; }
    .stSelectbox > div > div { background-color: #eceae4 !important; border-color: #c8c4bc !important; color: #1a1a1a !important; font-size: 1rem !important; }
    .stMultiSelect > div > div { background-color: #eceae4 !important; border-color: #c8c4bc !important; font-size: 1rem !important; }
    .stMultiSelect span { font-size: 0.95rem !important; }
    .stDownloadButton > button { background: #1a5fa0; color: #ffffff; border: 1px solid #1a5fa0; border-radius: 4px; font-family: 'IBM Plex Mono', monospace; font-size: 0.95rem; font-weight: 600; padding: 10px 20px; width: 100%; }
    .stDownloadButton > button:hover { background: #2a6fb0; }
    .stDataFrame { border: 1px solid #c8c4bc; border-radius: 6px; }
    .stAlert { border-radius: 4px; font-size: 1rem !important; }
    .drill-box { background: #eceae4; border: 1px solid #1a5fa0; border-radius: 6px; padding: 18px 22px; margin-top: 12px; }
    .drill-label { font-family: IBM Plex Mono, monospace; font-size: 0.8rem; color: #1a5fa0; letter-spacing: 0.15em; text-transform: uppercase; font-weight: 700; margin-bottom: 10px; }
    .drill-hint { font-size: 0.9rem; color: #666666; margin-bottom: 8px; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
TEMP_PATH_KEY = "temp_excel_path"
ORIG_NAME_KEY = "original_filename"
LANG_KEY      = "language"
DRILL_KEY     = "drill_data"   # stores current drill-down dataframe

SEASONS = ["S26", "S25", "S24", "S23", "Y22", "Y21", "Y20", "Y19", "Y18"]

ALL_PRODUCTS = [
    "SBM 48", "SBM 46", "SBO", "SBH", "Corn", "Wheat",
    "Rapeseed", "Barley", "SB", "SFO", "SFMG", "SBC",
    "Pellet", "Straw pellets", "Peas",
]

TRADER_MAP = {"TK": "Tomasz Koziarz", "NH": "Nick Halliwell"}

# Polish product name → short label (Wagi total товар column)
WAGI_PRODUCT_MAP = {
    "Śruta poekstrakcyjna paszowa z nasion soi BEZ GMO 48%":                     "SBM 48",
    "Śruta poekstrakcyjna paszowa z nasion soi BEZ GMO 48%; Usługa transportowa": "SBM 48",
    "Śruta poekstrakcyjna paszowa z nasion soi BEZ GMO 46%":                     "SBM 46",
    "Śruta poekstrakcyjna paszowa z nasion soi BEZ GMO 46%; Usługa transportowa": "SBM 46",
    "Łuska sojowa granulowana, materiał paszowy, bez GMO":                        "SBH",
    "Łuska sojowa granulowana, materiał paszowy, bez GMO; Usługa transportowa":   "SBH",
    "Olej sojowy surowy odgumowany, materiał paszowy, bez GMO":                   "SBO",
    "Olej sojowy surowy odgumowany, bez GMO":                                     "SBO",
    "Ziarno soi, materiał paszowy, bez GMO":                                      "SB",
    "Ziarno soi, bez GMO":                                                        "SB",
    "Kukurydza paszowa":                                                          "Corn",
    "Rzepak konsumpcyjny":                                                        "Rapeseed",
    "Pszenica paszowa":                                                           "Wheat",
    "Jęczmień paszowy":                                                           "Barley",
    "Soufflet - jęczmień browarny - Planet":                                      "Barley",
    "Soufflet - jęczmień browarny - Sebastian":                                   "Barley",
    "Pellet drzewny":                                                             "Pellet",
    "Pellet Paliwowy Pochodzenia Roślinnego z Soi":                               "Pellet",
    "Groch żółty łuskany (połówki)":                                              "Peas",
}

COUNTRY_MAP = {
    "Niemcy": "Germany", "PL": "Poland", "Polska": "Poland",
    "Czechy": "Czech Republic", "Słowacja": "Slovakia", "Austria": "Austria",
    "Dania": "Denmark", "Węgry": "Hungary", "Holandia": "Netherlands",
    "Chiny": "China", "Estonia": "Estonia", "Szwajcaria": "Switzerland",
    "Francja": "France", "Malezja": "Malaysia", "Szwecja": "Sweden",
    "DE": "Germany", "CZ": "Czech Republic", "SK": "Slovakia",
    "AT": "Austria", "DK": "Denmark", "HU": "Hungary",
    "NL": "Netherlands", "CN": "China", "EE": "Estonia",
    "CH": "Switzerland", "FR": "France", "MY": "Malaysia", "SE": "Sweden",
}

TRANSPORT_NORM = {
    "big-bag": "Big Bag", "big bag": "Big Bag", "big-bag/10USD/t": "Big Bag",
    "train": "Train", "train-M": "Train", "train/containers": "Train / Containers",
    "containers": "Containers",
    "vessel": "Vessel",
    "cars": "Truck", "cars UA": "Truck",
    "flexi": "Flexi Tank",
    "Medyka": "Medyka", "TR": "Transit", "Trans": "Transit",
}

PIE_COLORS = ["#3a7bd5","#e09030","#4caf70","#d04040","#a040d0",
              "#40c8d0","#d07040","#8090b0","#60c040","#d040a0"]

# ── Session state ─────────────────────────────────────────────────────────────
for k, v in [(TEMP_PATH_KEY, None), (ORIG_NAME_KEY, None), (LANG_KEY, "EN"), (DRILL_KEY, None)]:
    if k not in st.session_state:
        st.session_state[k] = v

# ── Helpers ───────────────────────────────────────────────────────────────────

@st.cache_data(ttl=300)
def load_sheet(path: str, sheet: str, nrows=None) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=sheet, nrows=nrows, engine="openpyxl")
    except Exception:
        return pd.DataFrame()


def _col(df, *names, fallback_idx=None):
    """Find a column by name (case-insensitive, stripped). Falls back to index if not found."""
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for name in names:
        found = lookup.get(str(name).strip().lower())
        if found is not None:
            return found
    if fallback_idx is not None and fallback_idx < len(df.columns):
        return df.columns[fallback_idx]
    return None


def get_download_bytes(path: str) -> bytes:
    with open(path, "rb") as f:
        return f.read()


def _build_buyer_excel(df_buyer):
    """Export the buyer summary table to Excel with totals row."""
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyer Summary"

    hdr_font  = Font(name="Arial", bold=True, size=10)
    hdr_fill  = PatternFill("solid", fgColor="DAEEF3")
    data_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft  = Alignment(horizontal="left",   vertical="center")
    rgt  = Alignment(horizontal="right",  vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Buyer", "Full Name", "Contracts",
               "Sold (MT)", "Issued (MT)", "Left (MT)",
               "Value → EUR", "Value → USD", "Value → PLN",
               "Avg Price EUR", "Avg Price USD", "Avg Price PLN"]
    col_widths = [18, 40, 11, 12, 12, 12, 16, 16, 16, 14, 14, 14]

    for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = ctr; c.border = bdr
        ws.column_dimensions[c.column_letter].width = w

    num_cols = set(range(3, len(headers) + 1))  # columns 3 onwards are numeric
    for ri, row in enumerate(df_buyer.itertuples(index=False), 2):
        vals = list(row)
        is_total = str(vals[0]).upper() == "TOTAL"
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = bold_font if is_total else data_font
            c.alignment = rgt if ci in num_cols else lft
            c.border = bdr
            if pd.notna(v) and ci in num_cols:
                c.number_format = '#,##0.00'

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_broker_excel(df_br, df_wagi, C):
    """Build broker commission Excel report matching the Agri Trade template layout."""
    # ── Monthly delivery pivot from Wagi (actual MT per contract per month) ──
    monthly_pivot = pd.DataFrame()
    all_months    = []
    if df_wagi is not None and isinstance(df_wagi, pd.DataFrame) and not df_wagi.empty:
        wm = df_wagi.copy()
        wm = wm[wm["Date_WZ"].notna()]
        wm["_ym"] = wm["Date_WZ"].dt.to_period("M")
        monthly_pivot = (
            wm.groupby(["_contract_key", "_ym"])["Qty_MT"]
            .sum().unstack(fill_value=0)
        )
        monthly_pivot.index = monthly_pivot.index.map(str)
        all_months = sorted(monthly_pivot.columns, key=lambda p: p.ordinal)
        # Only keep months that have deliveries for the contracts in this report
        keys_in_report = set(df_br["_contract_key"].astype(str))
        avail_keys = monthly_pivot.index.intersection(keys_in_report)
        if len(avail_keys) > 0:
            all_months = [p for p in all_months if monthly_pivot.loc[avail_keys, p].sum() > 0]
        else:
            all_months = []

    # ── Styles ──────────────────────────────────────────────────────────────
    hdr_font  = Font(name="Arial", bold=True, size=10)
    hdr_fill  = PatternFill("solid", fgColor="DAEEF3")
    data_font = Font(name="Arial", size=10)
    total_font= Font(name="Arial", bold=True, size=10)
    hdr_ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=False)
    rgt       = Alignment(horizontal="right",  vertical="center", wrap_text=False)
    lft       = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    thin      = Side(style="thin", color="AAAAAA")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    fixed_hdrs    = ["Date of contract", "Contract number", "Buyer", "Goods",
                     "Amount, mt", "Delivered, mt", "Left to deliver, mt"]
    trailing_hdrs = ["Broker", "Rate", "Currency", "Amount due, EUR"]
    month_labels  = [str(p.strftime("%b %y")) for p in all_months]
    all_hdrs      = fixed_hdrs + month_labels + trailing_hdrs

    # Column indices (1-based): weights (3dp), amount_due (2dp), rate (general)
    n_month_cols     = len(all_months)
    weight_cols      = set(range(5, 8)) | set(range(8, 8 + n_month_cols))
    rate_col         = 8 + n_month_cols + 1
    amt_due_col      = 8 + n_month_cols + 3
    numeric_cols     = weight_cols | {rate_col, amt_due_col}
    last_month_col   = 7 + n_month_cols if n_month_cols > 0 else None   # last delivery month col
    highlight_cols   = {amt_due_col} | ({last_month_col} if last_month_col else set())
    broker_col_idx   = 8 + n_month_cols                                  # left-align text
    left_align_cols  = {3, broker_col_idx}                               # Buyer=3, Broker

    for ci, hdr in enumerate(all_hdrs, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_ctr;  c.border = bdr

    # Strip trailing (Abbrev) bracket helper
    def _strip_bracket(s):
        return re.sub(r'\s*\([^)]*\)\s*$', '', str(s or "")).strip()

    # Running totals for the Total row
    tot_sold      = 0.0
    tot_deliv     = 0.0
    tot_left      = 0.0                        # only sums rows where left >= 9
    tot_months    = {p: 0.0 for p in all_months}
    tot_amt_due   = 0.0

    for ri, (_, row) in enumerate(df_br.iterrows(), 2):
        ckey     = str(row.get("_contract_key", ""))
        comm     = pd.to_numeric(
            str(row.get(C.get("commission",""), "") or "").replace(",", "."), errors="coerce"
        )
        sold     = pd.to_numeric(row.get(C["sold_mt"],    0), errors="coerce") or 0
        deliv_mt = pd.to_numeric(row.get("_delivery_mt",  0), errors="coerce") or 0
        amt_due  = round(float(row["_commission_eur"]), 2) if pd.notna(row.get("_commission_eur")) else None

        # Date as dd/mm/yyyy string only
        raw_date = row.get(C["date"])
        if pd.notna(raw_date):
            try:
                date_str = pd.Timestamp(raw_date).strftime("%d/%m/%Y")
            except Exception:
                date_str = str(raw_date)[:10]
        else:
            date_str = None

        buyer_clean  = _strip_bracket(row.get(C["buyer"], ""))
        broker_clean = _strip_bracket(row.get(C.get("broker", ""), ""))

        # Left to deliver: show "-" when value < 9 (completed / over-delivered)
        left_val     = float(sold) - float(deliv_mt)
        left_display = round(left_val, 3) if left_val >= 9 else "-"

        # Accumulate totals
        tot_sold   += float(sold)
        tot_deliv  += float(deliv_mt)
        if left_val >= 9:
            tot_left += round(left_val, 3)
        if amt_due is not None:
            tot_amt_due += amt_due

        values = [
            date_str,
            row.get(C["contract"], ""),
            buyer_clean,
            row.get(C["goods"], ""),
            round(float(sold), 3),
            round(float(deliv_mt), 3),
            left_display,
        ]
        for p in all_months:
            if not monthly_pivot.empty and ckey in monthly_pivot.index and p in monthly_pivot.columns:
                v = monthly_pivot.loc[ckey, p]
                mv = round(float(v), 3) if v else None
            else:
                mv = None
            values.append(mv)
            if mv is not None:
                tot_months[p] = tot_months.get(p, 0.0) + float(mv)
        values += [
            broker_clean,
            comm,
            str(row.get(C["currency"], "") or ""),
            amt_due,
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = data_font
            if ci in left_align_cols:
                c.alignment = lft
            elif ci in numeric_cols and val != "-":
                c.alignment = rgt
            else:
                c.alignment = ctr
            c.border = bdr
            if ci in highlight_cols:
                c.fill = hdr_fill
            if ci in weight_cols and isinstance(val, (int, float)):
                c.number_format = "0.000"
            elif ci == amt_due_col and val is not None:
                c.number_format = "0.00"

    # ── Total row ────────────────────────────────────────────────────────────
    total_row  = ri + 1 if df_br is not None and not df_br.empty else 2

    tot_values = [
        None,                        # Date
        None,                        # Contract
        None,                        # Buyer
        "Total",                     # Goods
        round(tot_sold,  3),         # Amount MT
        round(tot_deliv, 3),         # Delivered MT
        round(tot_left,  3),         # Left to deliver (numeric sum)
    ]
    for p in all_months:
        mv = tot_months.get(p, 0.0)
        tot_values.append(round(mv, 3) if mv else None)
    tot_values += [
        None,                        # Broker
        None,                        # Rate
        None,                        # Currency
        round(tot_amt_due, 2),       # Amount due
    ]
    for ci, val in enumerate(tot_values, 1):
        c = ws.cell(row=total_row, column=ci, value=val)
        c.font = total_font
        c.alignment = rgt if ci in numeric_cols else ctr
        c.border = bdr
        if ci in highlight_cols:
            c.fill = hdr_fill
        if ci in weight_cols and isinstance(val, (int, float)):
            c.number_format = "0.000"
        elif ci == amt_due_col and val is not None:
            c.number_format = "0.00"

    # ── Column widths (fixed per spec) ──────────────────────────────────────
    # openpyxl width=9 renders as ~8.22 in Excel due to padding; use 9.86 to display as 9
    # cols 1-2: 11, col 3 (Buyer): 30, cols 4-7+months: display-9, broker: 16, rate: 7,
    # currency: display-9, amount due: display-9
    W9 = 9.86   # value that displays as 9 in Excel
    fixed_col_widths = {1: 11, 2: 11, 3: 30}
    for ci in range(4, 8 + n_month_cols + 1):   # Goods → last month
        fixed_col_widths[ci] = W9
    fixed_col_widths[broker_col_idx]  = 16
    fixed_col_widths[rate_col]        = 7
    fixed_col_widths[rate_col + 1]    = W9   # Currency
    fixed_col_widths[amt_due_col]     = W9
    for ci in range(1, len(all_hdrs) + 1):
        w = fixed_col_widths.get(ci, W9)
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    ws.freeze_panes  = "A2"
    ws.row_dimensions[1].height = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_price_excel(df_chart, period_mode, currency, show_by_product):
    """
    Export the average FCA price chart data to Excel.
    df_chart  : DataFrame with columns [date, price, qty, product]
    period_mode: "Daily" | "Weekly" | "Monthly"
    currency  : "EUR" | "USD"
    show_by_product: bool
    """
    freq_map = {"Daily": "D", "Weekly": "W-MON", "Monthly": "MS"}
    freq = freq_map.get(period_mode, "MS")

    df = df_chart.copy()
    df["date"] = pd.to_datetime(df["date"])
    df["qty"]  = pd.to_numeric(df["qty"], errors="coerce").fillna(1)
    df["price"]= pd.to_numeric(df["price"], errors="coerce")
    df = df.dropna(subset=["price"])

    # Bin dates into periods
    df["period"] = df["date"].dt.to_period(
        "D" if freq == "D" else ("W" if freq == "W-MON" else "M")
    ).dt.to_timestamp()

    sym = "€" if currency == "EUR" else "$"
    col_label = f"Avg Price ({currency})"

    if show_by_product:
        grp = df.groupby(["period", "product"]).apply(
            lambda g: (g["price"] * g["qty"]).sum() / g["qty"].sum()
        ).reset_index(name="avg_price")
    else:
        grp = df.groupby("period").apply(
            lambda g: (g["price"] * g["qty"]).sum() / g["qty"].sum()
        ).reset_index(name="avg_price")

    grp = grp.sort_values("period")

    # ── Styles ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Price Data"

    hdr_font = Font(name="Arial", bold=True, size=10)
    hdr_fill = PatternFill("solid", fgColor="DAEEF3")
    dat_font = Font(name="Arial", size=10)
    ctr      = Alignment(horizontal="center", vertical="center")
    rgt      = Alignment(horizontal="right",  vertical="center")
    thin     = Side(style="thin", color="AAAAAA")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Format period label based on mode
    def fmt_period(ts):
        if period_mode == "Daily":
            return ts.strftime("%d/%m/%Y")
        elif period_mode == "Weekly":
            return f"W/c {ts.strftime('%d/%m/%Y')}"
        else:
            return ts.strftime("%b %Y")

    # Headers
    headers = (["Period", "Product", col_label] if show_by_product
               else ["Period", col_label])
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = ctr; c.border = bdr

    # Data rows
    for ri, (_, row) in enumerate(grp.iterrows(), 2):
        period_str = fmt_period(row["period"])
        avg        = round(float(row["avg_price"]), 2) if pd.notna(row["avg_price"]) else None
        if show_by_product:
            vals = [period_str, str(row.get("product", "")), avg]
        else:
            vals = [period_str, avg]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = dat_font
            c.border = bdr
            if isinstance(val, (int, float)):
                c.alignment = rgt
                c.number_format = f'"{sym}"#,##0.00'
            else:
                c.alignment = ctr

    # Column widths
    col_widths = [14, 16, 16] if show_by_product else [14, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@st.cache_data(ttl=3600)
def fetch_nbp_rates():
    rates = {}
    for code in ["USD", "EUR"]:
        try:
            url  = f"https://api.nbp.pl/api/exchangerates/rates/A/{code}/?format=json"
            data = requests.get(url, timeout=5).json()
            rates[code] = {"rate": data["rates"][0]["mid"], "date": data["rates"][0]["effectiveDate"]}
        except Exception:
            rates[code] = {"rate": None, "date": None}
    return rates


def sec(label):
    st.markdown(f"<div class='sec-hdr'>{label}</div>", unsafe_allow_html=True)


def sub(label):
    st.markdown(f"<div class='sub-hdr'>{label}</div>", unsafe_allow_html=True)


def tr(key):
    return T[st.session_state[LANG_KEY]].get(key, T["EN"].get(key, key))


def norm_key(x) -> str:
    """Normalise a contract key for matching (matches reference code logic)."""
    if pd.isna(x):
        return ""
    s = str(x)
    s = re.sub(r"[\x00-\x1f\x7f]", "", s)
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s


def silo_card(silo, goods, stocks, capacity, occ):
    pct       = min(max(occ * 100, 0), 100) if occ <= 1 else min(float(occ), 100)
    bar_color = "#c0392b" if pct > 80 else ("#d07020" if pct > 50 else "#1a5fa0")
    free      = max(float(capacity) - float(stocks), 0)
    return (
        f"<div style='background:#eceae4;border:1px solid #c8c4bc;border-radius:6px;"
        f"padding:18px 20px;margin-bottom:12px;'>"
        f"<div style='display:flex;justify-content:space-between;align-items:baseline;margin-bottom:10px;'>"
        f"<span style='font-family:IBM Plex Mono,monospace;font-size:1rem;color:#666666;"
        f"letter-spacing:0.12em;text-transform:uppercase;font-weight:700;'>SILO {silo}</span>"
        f"<span style='font-family:IBM Plex Mono,monospace;font-size:1rem;color:#2a2a2a;"
        f"font-weight:600;'>{goods}</span></div>"
        f"<div style='font-family:IBM Plex Mono,monospace;font-size:1.7rem;color:#1a1a1a;"
        f"font-weight:700;margin-bottom:10px;'>{stocks:,.2f} MT</div>"
        f"<div style='background:#d8d6d0;border-radius:3px;height:10px;margin-bottom:10px;'>"
        f"<div style='background:{bar_color};width:{pct:.1f}%;height:10px;border-radius:3px;'></div></div>"
        f"<div style='display:flex;justify-content:space-between;'>"
        f"<span style='font-family:IBM Plex Mono,monospace;font-size:0.9rem;color:#666666;"
        f"font-weight:600;'>{pct:.0f}% of {int(capacity):,} MT</span>"
        f"<span style='font-family:IBM Plex Mono,monospace;font-size:0.9rem;color:#666666;"
        f"font-weight:600;'>{tr('free')}: {free:,.2f} MT</span></div></div>"
    )


def pie_chart(labels, values, title):
    fig = go.Figure(go.Pie(
        labels=[str(l) for l in labels], values=[float(v) for v in values],
        hole=0.45,
        marker=dict(colors=PIE_COLORS[:len(labels)], line=dict(color="#f5f4f0", width=2)),
        textfont=dict(family="IBM Plex Mono", size=13, color="#1a1a1a"),
        hovertemplate="%{label}: %{value:,.2f} (%{percent})<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text=title, font=dict(family="IBM Plex Mono", size=14, color="#1a5fa0")),
        paper_bgcolor="#f5f4f0", plot_bgcolor="#f5f4f0",
        font=dict(family="IBM Plex Sans", color="#1a1a1a"),
        legend=dict(font=dict(family="IBM Plex Mono", size=12, color="#2a2a2a"),
                    bgcolor="#eceae4", bordercolor="#c8c4bc", borderwidth=1),
        margin=dict(t=50, b=20, l=10, r=10), height=340,
        # clickmode "event+select" makes slice clicks fire the plotly_selected event
        # which Streamlit's on_select listens to (it does NOT listen to plotly_click)
        clickmode="event+select",
    )
    return fig


def _pie_drilldown(fig, pie_labels, key, df_d, filter_col, group_col,
                   value_col, fmt_fn, bar_title_tpl):
    """
    Render a pie chart followed by a selectbox.
    When a category is chosen the selectbox shows a top-N bar chart below.
    """
    st.plotly_chart(fig, use_container_width=True,
                    config={"displayModeBar": False}, key=f"{key}_pie")

    chosen = st.selectbox(
        "🔍 Drill into",
        ["—"] + [str(l) for l in pie_labels],
        key=f"{key}_sel",
    )
    if chosen == "—":
        return

    df_d = df_d.copy()
    df_d[value_col] = pd.to_numeric(df_d[value_col], errors="coerce").fillna(0)
    if group_col in df_d.columns:
        df_d[group_col] = (df_d[group_col].astype(str).str.strip()
                           .replace({"": "Unknown", "nan": "Unknown"}))

    subset = df_d[df_d[filter_col].astype(str).str.strip() == chosen.strip()]
    if subset.empty:
        st.caption("No detail data for this category.")
        return

    contrib = (subset.groupby(group_col)[value_col].sum()
               .reset_index().rename(columns={value_col: "_v"}))
    contrib = (contrib[contrib["_v"] > 0]
               .sort_values("_v", ascending=False)
               .reset_index(drop=True))
    if contrib.empty:
        st.caption("No contributors found.")
        return

    max_n = len(contrib)
    opts  = sorted({n for n in [3, 5, 10] if n <= max_n} | {max_n})
    default_idx = min(len(opts) - 1, 2)
    top_n = st.radio("Show top", opts, index=default_idx,
                     horizontal=True, key=f"{key}_topn")

    top_rows  = contrib.head(top_n).copy()
    other_sum = contrib.iloc[top_n:]["_v"].sum()
    if other_sum > 0:
        top_rows = pd.concat(
            [top_rows, pd.DataFrame({group_col: ["Other"], "_v": [other_sum]})],
            ignore_index=True,
        )

    colors = [
        PIE_COLORS[i % len(PIE_COLORS)] if g != "Other" else "#b0aaa0"
        for i, g in enumerate(top_rows[group_col])
    ]
    bar = go.Figure(go.Bar(
        x=top_rows[group_col], y=top_rows["_v"],
        marker_color=colors,
        text=[fmt_fn(v) for v in top_rows["_v"]],
        textposition="outside",
        hovertemplate="%{x}: %{text}<extra></extra>",
    ))
    bar.update_layout(
        title=dict(text=bar_title_tpl.format(chosen),
                   font=dict(family="IBM Plex Mono", size=13, color="#1a5fa0")),
        paper_bgcolor="#f5f4f0", plot_bgcolor="#f5f4f0",
        font=dict(family="IBM Plex Sans", color="#1a1a1a"),
        xaxis=dict(tickfont=dict(family="IBM Plex Mono", size=11), showgrid=False),
        yaxis=dict(tickfont=dict(family="IBM Plex Mono", size=11), gridcolor="#e0deda"),
        margin=dict(t=50, b=10, l=10, r=30), height=320,
        showlegend=False,
    )
    st.plotly_chart(bar, use_container_width=True,
                    config={"displayModeBar": False}, key=f"{key}_bar")


def product_matches(cell_value: str, selected_products: list) -> bool:
    """Exact-token match: 'SBM 48' must not match 'SBM 48-PR'."""
    v = str(cell_value).strip().upper().replace(" ", "")
    for p in selected_products:
        pn = p.upper().replace(" ", "")
        # exact match first
        if v == pn:
            return True
        # also catch e.g. "SBM46" matching "SBM 46"
        if v == pn.replace("-", ""):
            return True
    return False


def norm_transport(val) -> str:
    v = str(val).strip()
    if v in ("nan", "", "None"):
        return "Truck"   # blank = delivered by truck
    return TRANSPORT_NORM.get(v, v)


def norm_country(val) -> str:
    v = str(val).strip()
    return COUNTRY_MAP.get(v, v if v not in ("nan", "", "None") else "Unknown")


def show_drill(df: pd.DataFrame, title: str):
    """Store a dataframe in session state so it renders as a drill-down panel."""
    st.session_state[DRILL_KEY] = {"df": df, "title": title}


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    _logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
    if os.path.exists(_logo_path):
        st.image(_logo_path, use_container_width=True)
    st.markdown(
        "<div style='font-family:IBM Plex Mono;font-size:1.15rem;font-weight:700;"
        "color:#1a1a1a;margin-bottom:16px;letter-spacing:0.04em;'>SP Dashboard</div>",
        unsafe_allow_html=True,
    )
    lang_choice = st.selectbox(
        tr("language"), options=["EN", "PL", "RU"],
        index=["EN", "PL", "RU"].index(st.session_state[LANG_KEY]),
        format_func=lambda x: {"EN": "🇬🇧 English", "PL": "🇵🇱 Polski", "RU": "🇷🇺 Русский"}[x],
        key="lang_select",
    )
    if lang_choice != st.session_state[LANG_KEY]:
        st.session_state[LANG_KEY] = lang_choice
        st.rerun()

    st.markdown("---")
    uploaded_file = st.file_uploader(tr("upload"), type=["xlsx"])
    if uploaded_file is not None:
        if st.session_state[ORIG_NAME_KEY] != uploaded_file.name:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.write(uploaded_file.read())
            tmp.close()
            old = st.session_state.get(TEMP_PATH_KEY)
            if old and os.path.isfile(old) and old != tmp.name:
                try: os.unlink(old)
                except Exception: pass
            st.session_state[TEMP_PATH_KEY] = tmp.name
            st.session_state[ORIG_NAME_KEY] = uploaded_file.name
            load_sheet.clear()
            st.session_state[DRILL_KEY] = None

    file_ok = (
        bool(st.session_state.get(TEMP_PATH_KEY))
        and os.path.isfile(st.session_state.get(TEMP_PATH_KEY) or "")
    )
    if file_ok:
        st.success(f"✓  {st.session_state[ORIG_NAME_KEY]}")
        st.markdown("---")
        st.markdown(
            f"<div style='font-size:0.95rem;color:#505050;font-weight:600;"
            f"letter-spacing:0.06em;text-transform:uppercase;margin-bottom:8px;'>"
            f"{tr('save')}</div>", unsafe_allow_html=True,
        )
        st.download_button(
            label=tr("download"),
            data=get_download_bytes(st.session_state[TEMP_PATH_KEY]),
            file_name=st.session_state[ORIG_NAME_KEY],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

if not file_ok:
    st.title(tr("title"))
    st.markdown(
        f"<div style='font-size:1.2rem;color:#3a3a3a;margin-top:16px;font-weight:500;'>"
        f"{tr('upload_prompt')}</div>", unsafe_allow_html=True,
    )
    st.stop()

excel_path = st.session_state[TEMP_PATH_KEY]

# ══════════════════════════════════════════════════════════════════════════════
# LOAD SP ONCE — shared by contract + wagi sections
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=300)
def load_sp(path):
    df = load_sheet(path, "SP")
    if df.empty:
        return df
    df.columns = [str(c) for c in df.columns]
    # Column lookup by name — robust to column shifts; index is fallback only
    C = {
        "season":    _col(df, "Season",             fallback_idx=0),
        "price_fca": _col(df, "Price FCA",          fallback_idx=1),
        "price_eur": _col(df, "Price to EUR",       fallback_idx=2),
        "sold_eur":  _col(df, "Amount on sold EUR", fallback_idx=5),
        "price_usd": _col(df, "Price to USD",       fallback_idx=7),
        "sold_usd":  _col(df, "Amount on sold USD", fallback_idx=10),
        "trader":    _col(df, "Trader",             fallback_idx=13),
        "date":      _col(df, "Date",               fallback_idx=14),
        "transport": _col(df, "Transport type",     fallback_idx=15),
        "goods":     _col(df, "Protein",            fallback_idx=16),
        "contract":  _col(df, "Contract",           fallback_idx=17),
        "buyer":     _col(df, "Buyer",              fallback_idx=18),
        "sold_mt":   _col(df, "Goods sold",         fallback_idx=19),
        "issued_mt": _col(df, "Goods issued",       fallback_idx=21),
        "left_mt":   _col(df, "Left to issue",      fallback_idx=22),
        "status":    _col(df, "Contr status",       fallback_idx=24),
        "fca1":      _col(df, "Price FCA.1",        fallback_idx=48),
        "ex_usd":    _col(df, "Ex rate USD",        fallback_idx=50),
        "ex_eur":    _col(df, "Ex rate EUR",        fallback_idx=51),
        "eurusd":    _col(df, "EUR/USD",            fallback_idx=52),
        "incoterms":   _col(df, "Incoterms",                    fallback_idx=62),
        "country":     _col(df, "Country",                      fallback_idx=69),
        "currency":    _col(df, "Currency",                     fallback_idx=72),
        "broker":      _col(df, "Broker"),
        "commission":  _col(df, "Commision", "Commission"),
    }
    # Normalise transport: blank → Truck
    df[C["transport"]] = df[C["transport"]].apply(
        lambda x: "Truck" if pd.isna(x) or str(x).strip() in ("", "nan", "None") else str(x).strip()
    )
    # Parse date
    df["_date"] = pd.to_datetime(df[C["date"]], errors="coerce")
    df["_year"]  = df["_date"].dt.year
    df["_month"] = df["_date"].dt.month
    # Normalise numeric
    for k in ["price_fca", "ex_usd", "ex_eur", "sold_mt", "issued_mt", "left_mt", "sold_eur", "sold_usd"]:
        df[C[k]] = pd.to_numeric(df[C[k]], errors="coerce")
    # Contract norm key
    df["_contract_key"] = df[C["contract"]].apply(norm_key)
    # Extract buyer abbreviation from "Full Name (Abbreviation)" format
    def _extract_abbrev(v):
        m = re.search(r'\(([^)]+)\)\s*$', str(v).strip())
        return m.group(1).strip() if m else str(v).strip()
    df["_buyer_abbrev"] = df[C["buyer"]].apply(_extract_abbrev) if C.get("buyer") else ""
    return df, C


@st.cache_data(ttl=300)
def load_wagi(path):
    df = load_sheet(path, "Wagi total")
    if df.empty:
        return df
    df = df.rename(columns={
        df.columns[0]:  "Date_WZ",
        df.columns[5]:  "Product_PL",
        df.columns[9]:  "Contract_raw",   # № контракта
        df.columns[15]: "Qty_kg",
        df.columns[20]: "Season_col",
    })
    df["Date_WZ"]  = pd.to_datetime(df["Date_WZ"], errors="coerce")
    df["Qty_kg"]   = pd.to_numeric(df["Qty_kg"], errors="coerce").fillna(0)
    df["Qty_MT"]   = df["Qty_kg"] / 1000.0
    df["Product"]  = df["Product_PL"].map(WAGI_PRODUCT_MAP).fillna(df["Product_PL"].astype(str))
    df["_year"]    = df["Date_WZ"].dt.year
    df["_month"]   = df["Date_WZ"].dt.month
    df["_contract_key"] = df["Contract_raw"].apply(norm_key)
    return df


sp_result = load_sp(excel_path)
if isinstance(sp_result, tuple):
    df_sp_full, C = sp_result
else:
    df_sp_full, C = pd.DataFrame(), {}

# Derive available seasons/products from the actual file — new entries appear automatically
if not df_sp_full.empty and C:
    _sv = df_sp_full[C["season"]].dropna().astype(str).str.strip()
    avail_seasons = sorted(
        [v for v in _sv.unique() if v and v not in ("nan", "None", "")], reverse=True
    ) or SEASONS
    _pv = df_sp_full[C["goods"]].dropna().astype(str).str.strip()
    avail_products = sorted(
        [v for v in _pv.unique() if v and v not in ("nan", "None", "")]
    ) or ALL_PRODUCTS
else:
    avail_seasons  = SEASONS
    avail_products = ALL_PRODUCTS

# Buyer abbreviations and broker options for filters
if not df_sp_full.empty and C:
    if C.get("buyer"):
        avail_buyers = sorted({
            v for v in df_sp_full["_buyer_abbrev"].dropna().astype(str).str.strip()
            if v and v not in ("nan", "None", "")
        })
    else:
        avail_buyers = []
    if C.get("broker"):
        avail_brokers = sorted({
            v for v in df_sp_full[C["broker"]].dropna().astype(str).str.strip()
            if v and v not in ("nan", "None", "")
        })
    else:
        avail_brokers = []
else:
    avail_buyers  = []
    avail_brokers = []

df_wagi_full = load_wagi(excel_path)

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD TITLE
# ══════════════════════════════════════════════════════════════════════════════
st.title(tr("title"))
st.markdown(
    f"<div style='font-family:IBM Plex Mono;font-size:1rem;color:#666666;"
    f"font-weight:600;margin-top:-6px;margin-bottom:28px;'>"
    f"{tr('updated')}: {datetime.now().strftime('%d %b %Y  %H:%M')}</div>",
    unsafe_allow_html=True,
)

# ══════════════════════════════════════════════════════════════════════════════
# FILTERS — applied to EVERYTHING
# ══════════════════════════════════════════════════════════════════════════════
sec(tr("filters"))
fc1, fc2, fc3, fc4 = st.columns(4)

with fc1:
    _season_default = [avail_seasons[0]] if avail_seasons else ["All"]
    sel_seasons = st.multiselect(tr("season"), ["All"] + avail_seasons,
                                 default=_season_default, key="f_seasons")
    if "All" in sel_seasons or not sel_seasons:
        sel_seasons = avail_seasons

with fc2:
    sel_products = st.multiselect(tr("product"), ["All"] + avail_products, default=["All"], key="f_products")
    if "All" in sel_products or not sel_products:
        sel_products = avail_products
        all_products_selected = True
    else:
        all_products_selected = False

with fc3:
    trader_opts = ["All"] + list(TRADER_MAP.values()) + ["Other"]
    sel_trader_labels = st.multiselect(tr("trader"), trader_opts, default=["All"], key="f_traders")
    if "All" in sel_trader_labels or not sel_trader_labels:
        sel_trader_codes = list(TRADER_MAP.keys()) + ["__other__"]
    else:
        sel_trader_codes = []
        for lbl in sel_trader_labels:
            for code, name in TRADER_MAP.items():
                if name == lbl: sel_trader_codes.append(code)
            if lbl == "Other": sel_trader_codes.append("__other__")

with fc4:
    show_sections = st.multiselect(
        tr("show_sections"), tr("sections"), default=tr("sections"), key="f_sections"
    )
    section_en = {v: k for lang in T.values()
                  for k, v in zip(
                      ["Exchange Rates","Silo Occupancy","Contract Summary",
                       "Price Analysis","Unloading Schedule","Unloaded (Wagi Total)",
                       "Broker Report","Buyer Summary"],
                      lang.get("sections", [])
                  )}
    active_sections = {section_en.get(s, s) for s in show_sections}

dy1, dy2 = st.columns(2)
year_options = list(range(2021, datetime.now().year + 1))
with dy1:
    sel_years = st.multiselect(tr("year"), ["All"] + [str(y) for y in year_options], default=["All"], key="f_years")
    sel_years_int = year_options if ("All" in sel_years or not sel_years) else [int(y) for y in sel_years if y != "All"]

with dy2:
    month_names = tr("month_names")
    sel_months = st.multiselect(tr("month"), ["All"] + month_names, default=["All"], key="f_months")
    if "All" in sel_months or not sel_months:
        sel_month_nums = list(range(1, 13))
    else:
        EN_M = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        sel_month_nums = [i+1 for i, m in enumerate(month_names) if m in sel_months]

# Country filter row
all_country_english = sorted([v for v in set(COUNTRY_MAP.values()) if v != "Unknown"] + ["Poland"])
cy1, cy2 = st.columns([2, 2])
with cy1:
    sel_countries = st.multiselect(
        "Country", ["All"] + all_country_english, default=["All"], key="f_countries"
    )
    if "All" in sel_countries or not sel_countries:
        sel_countries = all_country_english + ["Unknown"]
    # Build the set of raw country values (as stored in SP) that match the selection
    sel_country_raw = set()
    for raw_pl, eng in COUNTRY_MAP.items():
        if eng in sel_countries:
            sel_country_raw.add(raw_pl)
    # "PL" and "Polska" map to Poland — only add if Poland is selected
    if "Poland" in sel_countries:
        sel_country_raw.add("PL")
        sel_country_raw.add("Polska")
    # Also add any English names directly (in case stored as English in file)
    for ec in sel_countries:
        sel_country_raw.add(ec)

bb1, bb2 = st.columns(2)
with bb1:
    sel_buyers = st.multiselect(
        tr("buyer") if "buyer" in T.get(st.session_state.get("lang","EN"),{}) else "Buyer",
        ["All"] + avail_buyers, default=["All"], key="f_buyers"
    )
    if "All" in sel_buyers or not sel_buyers:
        sel_buyers_set      = set(avail_buyers)
        all_buyers_selected = True
    else:
        sel_buyers_set      = set(sel_buyers)
        all_buyers_selected = False

with bb2:
    sel_brokers = st.multiselect(
        tr("broker") if "broker" in T.get(st.session_state.get("lang","EN"),{}) else "Broker",
        ["All"] + avail_brokers, default=["All"], key="f_brokers"
    )
    if "All" in sel_brokers or not sel_brokers:
        sel_brokers_set      = set(avail_brokers)
        all_brokers_selected = True
    else:
        sel_brokers_set      = set(sel_brokers)
        all_brokers_selected = False

st.markdown("<br>", unsafe_allow_html=True)

# ── Shared filter function for SP ─────────────────────────────────────────────
def apply_sp_filters(df):
    if df.empty or not C:
        return df
    # Season
    df = df[df[C["season"]].astype(str).isin(sel_seasons)]
    # Year + month on contract date.
    # NaT rows (unparseable dates e.g. "hg") only pass through when no date filter is active.
    no_date_filter = (set(sel_years_int) >= set(year_options)) and (sel_month_nums == list(range(1, 13)))
    date_mask = (df["_date"].isna() if no_date_filter else pd.Series(False, index=df.index)) | (
        df["_year"].isin(sel_years_int) & df["_month"].isin(sel_month_nums)
    )
    df = df[date_mask]
    # Product (exact token match)
    df = df[df[C["goods"]].apply(lambda x: product_matches(str(x), sel_products))]
    # Trader
    def trader_ok(v):
        v = str(v).strip()
        if "__other__" in sel_trader_codes and v not in TRADER_MAP: return True
        return v in sel_trader_codes
    df = df[df[C["trader"]].apply(trader_ok)]
    # Country — only skip filter if All countries are selected
    _all_countries_selected = len(sel_countries) >= len(all_country_english)
    def country_ok(v):
        if _all_countries_selected:
            return True
        v = str(v).strip()
        if not v or v in ("nan","None"):
            return False  # blank country: exclude when filtering to specific countries
        if v in sel_country_raw:
            return True
        eng = COUNTRY_MAP.get(v, v)
        return eng in sel_countries
    df = df[df[C["country"]].apply(country_ok)]
    # Buyer (by abbreviation extracted from "Full Name (Abbrev)" format)
    if not all_buyers_selected and C.get("buyer") and avail_buyers:
        df = df[df["_buyer_abbrev"].isin(sel_buyers_set)]
    # Broker
    if not all_brokers_selected and C.get("broker") and avail_brokers:
        df = df[df[C["broker"]].astype(str).str.strip().isin(sel_brokers_set)]
    return df


def apply_wagi_filters(df):
    if df.empty:
        return df
    # Season
    df = df[df["Season_col"].astype(str).isin(sel_seasons)]
    # Year + month
    df = df[df["_year"].isin(sel_years_int) & df["_month"].isin(sel_month_nums)]
    # Product
    df = df[df["Product"].apply(lambda x: product_matches(str(x), sel_products))]
    return df


df_sp  = apply_sp_filters(df_sp_full.copy()) if not df_sp_full.empty else pd.DataFrame()
df_w   = apply_wagi_filters(df_wagi_full.copy()) if not df_wagi_full.empty else pd.DataFrame()

# Propagate SP-only filters (trader / buyer / broker) to Wagi via contract key.
# These attributes live only in SP, so we apply them by restricting wagi to
# contract keys that survive the attribute filters on the full SP dataset.
if C and not df_sp_full.empty and not df_w.empty and "_contract_key" in df_w.columns:
    _sp_attr = df_sp_full.copy()
    # Trader
    if C.get("trader"):
        _sp_attr = _sp_attr[_sp_attr[C["trader"]].astype(str).str.strip().apply(
            lambda v: ("__other__" in sel_trader_codes and v not in TRADER_MAP)
                      or v in sel_trader_codes
        )]
    # Buyer
    if not all_buyers_selected and C.get("buyer") and avail_buyers:
        _sp_attr = _sp_attr[_sp_attr["_buyer_abbrev"].isin(sel_buyers_set)]
    # Broker
    if not all_brokers_selected and C.get("broker") and avail_brokers:
        _sp_attr = _sp_attr[_sp_attr[C["broker"]].astype(str).str.strip().isin(sel_brokers_set)]
    # Only restrict wagi if the attribute filters actually removed something
    if len(_sp_attr) < len(df_sp_full):
        _valid_wagi_keys = set(_sp_attr["_contract_key"].dropna().astype(str))
        df_w = df_w[
            df_w["_contract_key"].isna() |
            df_w["_contract_key"].astype(str).isin(_valid_wagi_keys)
        ]

# ── Drill-down panel ──────────────────────────────────────────────────────────
drill = st.session_state.get(DRILL_KEY)
if drill is not None:
    with st.container():
        st.markdown(f"<div class='drill-label'>{tr('drill_title')} — {drill['title']}</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='drill-hint'>{len(drill['df'])} rows</div>", unsafe_allow_html=True)
        st.dataframe(drill["df"], use_container_width=True, hide_index=True, height=400)
        if st.button(tr("drill_close"), key="drill_close_btn"):
            st.session_state[DRILL_KEY] = None
            st.rerun()
    st.markdown("---")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — EXCHANGE RATES
# ══════════════════════════════════════════════════════════════════════════════
if "Exchange Rates" in active_sections:
    sec(tr("exchange_rates"))
    rates    = fetch_nbp_rates()
    usd_rate = rates.get("USD", {}).get("rate")
    eur_rate = rates.get("EUR", {}).get("rate")
    eur_usd  = round(eur_rate / usd_rate, 4) if (usd_rate and eur_rate) else None
    r1, r2, r3, r4 = st.columns(4)
    with r1: st.metric("USD / PLN", f"{usd_rate:.4f}" if usd_rate else "N/A")
    with r2: st.metric("EUR / PLN", f"{eur_rate:.4f}" if eur_rate else "N/A")
    with r3: st.metric("EUR / USD", f"{eur_usd:.4f}" if eur_usd else "N/A")
    with r4: st.metric(tr("rate_date"), rates.get("USD", {}).get("date") or "—")
    st.markdown("<br>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — SILO OCCUPANCY (not filtered by year/month — always live)
# ══════════════════════════════════════════════════════════════════════════════
if "Silo Occupancy" in active_sections:
    sec(tr("silo_occ"))
    try:
        def parse_silos(sheet, nrows=50):
            df_raw = load_sheet(excel_path, sheet, nrows=nrows)
            result = []
            if df_raw.empty: return result
            for i in range(2, len(df_raw)):
                row = df_raw.iloc[i]
                try:
                    sno = row.iloc[13]; stk = row.iloc[16]; occ = row.iloc[17]
                    gds = row.iloc[18]; cap = row.iloc[19]
                    if pd.isna(sno) or str(sno).strip() in ("", "nan", "None", "Total", "TOTAL"):
                        continue
                    if pd.notna(stk):
                        result.append({
                            "Silo": int(sno) if str(sno).replace(".","").isdigit() else sno,
                            "Goods": str(gds).strip() if pd.notna(gds) else "",
                            "Stocks": round(float(stk), 1),
                            "Capacity": int(float(cap)) if pd.notna(cap) else 2000,
                            "Occ": float(occ) if pd.notna(occ) else 0,
                        })
                except Exception: continue
            return result

        meal_silos = parse_silos("Meals")
        oil_silos  = parse_silos("Oils")

        df_beans = load_sheet(excel_path, "Beans", nrows=80)
        beans_silos = []
        if not df_beans.empty:
            try:
                # Find data start: first row where col 1 is a valid integer silo number
                DATA_START_IDX = 21  # known fallback
                for _ti in range(5, len(df_beans)):
                    try:
                        int(float(df_beans.iloc[_ti].iloc[1]))
                        DATA_START_IDX = _ti
                        break
                    except (ValueError, TypeError):
                        continue
                for i in range(DATA_START_IDX, len(df_beans)):
                    row = df_beans.iloc[i]
                    try:
                        silo_no  = row.iloc[1]
                        stocks   = row.iloc[4]
                        occ      = row.iloc[5]
                        goods    = str(row.iloc[7]).strip()
                        capacity = row.iloc[8]
                        if pd.isna(silo_no) or str(silo_no).strip() in ("", "nan", "Total", "TOTAL"):
                            continue
                        if pd.notna(stocks):
                            try:
                                beans_silos.append({
                                    "Silo":     int(float(silo_no)),
                                    "Goods":    goods,
                                    "Stocks":   round(float(stocks), 1),
                                    "Capacity": int(float(capacity)) if pd.notna(capacity) else 7000,
                                    "Occ":      float(occ) if pd.notna(occ) else 0,
                                })
                            except Exception:
                                pass
                    except Exception:
                        continue
            except Exception:
                pass
        # Group ALL silos by their current goods content — any product in any silo is handled
        from collections import defaultdict
        _silo_groups = defaultdict(list)
        for _s in meal_silos + oil_silos + beans_silos:
            _g = str(_s["Goods"]).strip()
            if _g and _g not in ("nan", "None", ""):
                _silo_groups[_g].append(_s)

        def render_silo_group(label, silos):
            sub(label)
            for s in silos:
                st.markdown(silo_card(s["Silo"], s["Goods"], s["Stocks"],
                                      s["Capacity"], s["Occ"]), unsafe_allow_html=True)
            total = sum(s["Stocks"] for s in silos)
            st.markdown(f"<div style='font-family:IBM Plex Mono;font-size:1rem;color:#666666;"
                        f"font-weight:700;text-align:right;margin-top:4px;'>"
                        f"{tr('total')}: {total:,.2f} MT</div>", unsafe_allow_html=True)

        def _silo_visible(goods_label):
            if all_products_selected:
                return True
            if product_matches(goods_label, sel_products):
                return True
            short = WAGI_PRODUCT_MAP.get(goods_label)
            return bool(short and short in sel_products)

        _groups_to_render = [(g, sls) for g, sls in sorted(_silo_groups.items()) if _silo_visible(g)]

        if _groups_to_render:
            for _ci in range(0, len(_groups_to_render), 3):
                _chunk = _groups_to_render[_ci:_ci + 3]
                _cols = st.columns(len(_chunk))
                for _col, (_lbl, _sls) in zip(_cols, _chunk):
                    with _col:
                        render_silo_group(_lbl, _sls)
                st.markdown("<br>", unsafe_allow_html=True)
        else:
            st.info(tr("no_silo"))
    except Exception as e:
        st.warning(f"Could not load silo data: {e}")
    st.markdown("<br>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — CONTRACT SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
if "Contract Summary" in active_sections:
    season_label = " + ".join(sel_seasons) if len(sel_seasons) <= 4 else f"{len(sel_seasons)} seasons"
    sec(f"{tr('contract_summary')}  ·  {season_label}")

    if not df_sp.empty and C:
        try:
            df_c = df_sp.copy()

            # ── Avg Price FCA (quantity-weighted, per contract row) ───────────
            def wavg_fca(grp):
                w  = grp[C["sold_mt"]].fillna(0)
                p  = pd.to_numeric(grp[C["fca1"]], errors="coerce").fillna(0)
                tw = w.sum()
                return round((p * w).sum() / tw, 2) if tw > 0 else None

            # ── Per-currency value totals (by contract currency, no conversion) ─
            def currency_total(grp, cur):
                mask = grp[C["currency"]].astype(str).str.strip().str.upper() == cur
                sub_g = grp[mask]
                price = pd.to_numeric(sub_g[C["fca1"]], errors="coerce").fillna(0)
                qty   = sub_g[C["sold_mt"]].fillna(0)
                return (price * qty).sum()

            # ── Per-row converted values (all currencies → EUR / USD / PLN) ──
            def _to_eur(row):
                p   = pd.to_numeric(row.get(C["fca1"]),   errors="coerce")
                q   = pd.to_numeric(row.get(C["sold_mt"]), errors="coerce")
                cur = str(row.get(C["currency"], "")).strip().upper()
                if pd.isna(p) or pd.isna(q) or q == 0: return 0.0
                if cur == "EUR": return p * q
                ex_eur = pd.to_numeric(row.get(C["ex_eur"]), errors="coerce")
                eurusd = pd.to_numeric(row.get(C["eurusd"]), errors="coerce")
                if cur == "PLN" and pd.notna(ex_eur) and ex_eur != 0: return (p / ex_eur) * q
                if cur == "USD" and pd.notna(eurusd)  and eurusd != 0: return (p / eurusd)  * q
                return 0.0

            def _to_usd(row):
                p   = pd.to_numeric(row.get(C["fca1"]),   errors="coerce")
                q   = pd.to_numeric(row.get(C["sold_mt"]), errors="coerce")
                cur = str(row.get(C["currency"], "")).strip().upper()
                if pd.isna(p) or pd.isna(q) or q == 0: return 0.0
                if cur == "USD": return p * q
                ex_usd = pd.to_numeric(row.get(C["ex_usd"]), errors="coerce")
                eurusd = pd.to_numeric(row.get(C["eurusd"]), errors="coerce")
                if cur == "PLN" and pd.notna(ex_usd) and ex_usd != 0: return (p / ex_usd)  * q
                if cur == "EUR" and pd.notna(eurusd):                   return p * eurusd   * q
                return 0.0

            def _to_pln(row):
                p   = pd.to_numeric(row.get(C["fca1"]),   errors="coerce")
                q   = pd.to_numeric(row.get(C["sold_mt"]), errors="coerce")
                cur = str(row.get(C["currency"], "")).strip().upper()
                if pd.isna(p) or pd.isna(q) or q == 0: return 0.0
                if cur == "PLN": return p * q
                ex_eur = pd.to_numeric(row.get(C["ex_eur"]), errors="coerce")
                ex_usd = pd.to_numeric(row.get(C["ex_usd"]), errors="coerce")
                if cur == "EUR" and pd.notna(ex_eur): return p * ex_eur * q
                if cur == "USD" and pd.notna(ex_usd): return p * ex_usd * q
                return 0.0

            df_c["_row_eur"] = df_c.apply(_to_eur, axis=1)
            df_c["_row_usd"] = df_c.apply(_to_usd, axis=1)
            df_c["_row_pln"] = df_c.apply(_to_pln, axis=1)

            summary = df_c.groupby(C["goods"]).apply(lambda g: pd.Series({
                tr("contracts"):           len(g),
                tr("sold_mt"):             g[C["sold_mt"]].sum(),
                tr("issued_mt"):           g[C["issued_mt"]].sum(),
                tr("left_mt"):             g[C["left_mt"]].sum(),
                tr("avg_price"):           wavg_fca(g),
                "_val_eur":                currency_total(g, "EUR"),
                "_val_usd":                currency_total(g, "USD"),
                "_val_pln":                currency_total(g, "PLN"),
                "_conv_eur":               g["_row_eur"].sum(),
                "_conv_usd":               g["_row_usd"].sum(),
                "_conv_pln":               g["_row_pln"].sum(),
            })).reset_index().rename(columns={C["goods"]: tr("product")})

            summary = summary[summary[tr("sold_mt")] > 0].sort_values(tr("sold_mt"), ascending=False)

            total_eur = summary["_val_eur"].sum()
            total_usd = summary["_val_usd"].sum()
            total_pln = summary["_val_pln"].sum()
            total_conv_eur = summary["_conv_eur"].sum()
            total_conv_usd = summary["_conv_usd"].sum()
            total_conv_pln = summary["_conv_pln"].sum()

            # MT metrics — clicking shows drill-down
            m1, m2, m3, m4 = st.columns(4)
            with m1:
                st.metric(tr("contracts"), len(df_c))
                if st.button("🔍", key="drill_contracts"):
                    drill_df = df_c[[C["season"], C["date"], C["contract"], C["goods"],
                                     C["buyer"], C["trader"], C["sold_mt"], C["fca1"],
                                     C["currency"], C["transport"], C["country"], C["status"]]].copy()
                    drill_df.columns = ["Season","Date","Contract","Product","Buyer",
                                        "Trader","Sold MT","Price FCA","Currency",
                                        "Transport","Country","Status"]
                    drill_df["Date"] = pd.to_datetime(drill_df["Date"], errors="coerce").dt.strftime("%d/%m/%Y")
                    show_drill(drill_df, f"{tr('contracts')} ({len(drill_df)})")
                    st.rerun()
            with m2:
                st.metric(tr("sold_mt"), f"{summary[tr('sold_mt')].sum():,.2f}")
                if st.button("🔍", key="drill_sold"):
                    show_drill(df_c[[C["season"],C["contract"],C["goods"],C["sold_mt"],
                                     C["fca1"],C["currency"],C["trader"]]].rename(
                        columns={C["season"]:"Season",C["contract"]:"Contract",C["goods"]:"Product",
                                 C["sold_mt"]:"Sold MT",C["fca1"]:"Price FCA",
                                 C["currency"]:"Currency",C["trader"]:"Trader"}),
                        f"{tr('sold_mt')}")
                    st.rerun()
            with m3:
                st.metric(tr("issued_mt"), f"{summary[tr('issued_mt')].sum():,.2f}")
            with m4:
                st.metric(tr("left_mt"), f"{summary[tr('left_mt')].sum():,.2f}")

            st.markdown("<br>", unsafe_allow_html=True)
            sub(tr("value_sold"))

            # Row 1: value by contract currency (no conversion)
            st.caption("By contract currency")
            v1, v2, v3 = st.columns(3)
            with v1: st.metric(tr("sold_eur"), f"€ {total_eur:,.2f}" if total_eur else "—")
            with v2: st.metric(tr("sold_usd"), f"$ {total_usd:,.2f}" if total_usd else "—")
            with v3: st.metric(tr("sold_pln"), f"zł {total_pln:,.2f}" if total_pln else "—",
                                help=tr("pln_help"))

            st.markdown("<br>", unsafe_allow_html=True)

            # Row 2: everything converted to each currency
            st.caption("All contracts converted")
            c1, c2, c3 = st.columns(3)
            with c1: st.metric("Total → EUR", f"€ {total_conv_eur:,.2f}" if total_conv_eur else "—")
            with c2: st.metric("Total → USD", f"$ {total_conv_usd:,.2f}" if total_conv_usd else "—")
            with c3: st.metric("Total → PLN", f"zł {total_conv_pln:,.2f}" if total_conv_pln else "—")

            st.markdown("<br>", unsafe_allow_html=True)

            # Table
            disp = summary.drop(columns=["_val_eur","_val_usd","_val_pln",
                                          "_conv_eur","_conv_usd","_conv_pln"]).copy()
            # Total row (before string formatting)
            total_row_cs = {
                tr("product"):   "TOTAL",
                tr("contracts"): int(disp[tr("contracts")].sum()),
                tr("sold_mt"):   disp[tr("sold_mt")].sum(),
                tr("issued_mt"): disp[tr("issued_mt")].sum(),
                tr("left_mt"):   disp[tr("left_mt")].sum(),
                tr("avg_price"): None,
            }
            disp = pd.concat([disp, pd.DataFrame([total_row_cs])], ignore_index=True)
            disp[tr("avg_price")]  = disp[tr("avg_price")].apply(lambda x: f"{x:,.2f}" if pd.notna(x) and x else "—")
            disp[tr("sold_mt")]    = disp[tr("sold_mt")].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "—")
            disp[tr("issued_mt")]  = disp[tr("issued_mt")].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "—")
            disp[tr("left_mt")]    = disp[tr("left_mt")].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "—")
            disp[tr("contracts")]  = disp[tr("contracts")].apply(lambda x: int(x) if pd.notna(x) else "")
            st.dataframe(disp, use_container_width=True, hide_index=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Pre-compute normalised columns used for drill-down filtering
            df_c["_norm_transport"] = df_c[C["transport"]].apply(lambda x: norm_transport(str(x)))
            df_c["_norm_country"]   = df_c[C["country"]].apply(lambda x: norm_country(str(x)))
            df_c[C["sold_mt"]]      = pd.to_numeric(df_c[C["sold_mt"]], errors="coerce").fillna(0)

            # ── Product pie charts (Revenue + Tonnes) ────────────────────────
            st.markdown("<br>", unsafe_allow_html=True)
            pie_cur = st.radio(
                "Display currency", ["EUR", "USD", "PLN"],
                horizontal=True, key="pie_currency"
            )
            _cur_col      = {"EUR": "_conv_eur",  "USD": "_conv_usd",  "PLN": "_conv_pln"}[pie_cur]
            _drill_val_col = {"EUR": "_row_eur",   "USD": "_row_usd",   "PLN": "_row_pln"}[pie_cur]
            _cur_sym      = {"EUR": "€",           "USD": "$",          "PLN": "zł"}[pie_cur]

            pp1, pp2 = st.columns(2)
            with pp1:
                rev_data = summary[summary[_cur_col] > 0].copy()
                if not rev_data.empty:
                    _rev_labels = rev_data[tr("product")].tolist()
                    _pie_drilldown(
                        pie_chart(_rev_labels, rev_data[_cur_col].tolist(),
                                  f"Revenue by product ({_cur_sym})"),
                        pie_labels=_rev_labels,
                        key="pie_rev",
                        df_d=df_c, filter_col=C["goods"], group_col="_buyer_abbrev",
                        value_col=_drill_val_col,
                        fmt_fn=lambda v: f"{_cur_sym}{v:,.2f}",
                        bar_title_tpl="Top buyers — {{}} · revenue ({})".format(_cur_sym),
                    )
            with pp2:
                mt_data = summary[summary[tr("sold_mt")] > 0].copy()
                if not mt_data.empty:
                    _mt_labels = mt_data[tr("product")].tolist()
                    _pie_drilldown(
                        pie_chart(_mt_labels, mt_data[tr("sold_mt")].tolist(),
                                  "Tonnes sold by product"),
                        pie_labels=_mt_labels,
                        key="pie_mt",
                        df_d=df_c, filter_col=C["goods"], group_col="_buyer_abbrev",
                        value_col=C["sold_mt"],
                        fmt_fn=lambda v: f"{v:,.2f} MT",
                        bar_title_tpl="Top buyers — {} · tonnes",
                    )

            # ── Transport / Country breakdown ─────────────────────────────────
            st.markdown("<br>", unsafe_allow_html=True)

            # Build aggregated transport and country dataframes
            tg = (df_c.groupby("_norm_transport")[C["sold_mt"]].sum()
                  .reset_index().rename(columns={"_norm_transport": "_label"}))
            tg = tg[tg[C["sold_mt"]] > 0].sort_values(C["sold_mt"], ascending=False)

            cg = df_c[df_c["_norm_country"] != "Unknown"].copy()
            cg = (cg.groupby("_norm_country")[C["sold_mt"]].sum()
                  .reset_index().rename(columns={"_norm_country": "_label"}))
            cg = cg[cg[C["sold_mt"]] > 0].sort_values(C["sold_mt"], ascending=False)

            p1, p2 = st.columns(2)
            with p1:
                if not tg.empty:
                    _tg_labels = tg["_label"].tolist()
                    _pie_drilldown(
                        pie_chart(_tg_labels, tg[C["sold_mt"]].tolist(),
                                  tr("transport_chart")),
                        pie_labels=_tg_labels,
                        key="pie_transport",
                        df_d=df_c, filter_col="_norm_transport", group_col="_buyer_abbrev",
                        value_col=C["sold_mt"],
                        fmt_fn=lambda v: f"{v:,.2f} MT",
                        bar_title_tpl="Top buyers — {} · transport",
                    )
            with p2:
                if not cg.empty:
                    _cg_labels = cg["_label"].tolist()
                    _pie_drilldown(
                        pie_chart(_cg_labels, cg[C["sold_mt"]].tolist(),
                                  tr("country_chart")),
                        pie_labels=_cg_labels,
                        key="pie_country",
                        df_d=df_c, filter_col="_norm_country", group_col="_buyer_abbrev",
                        value_col=C["sold_mt"],
                        fmt_fn=lambda v: f"{v:,.2f} MT",
                        bar_title_tpl="Top buyers — {}",
                    )
        except Exception as e:
            st.warning(f"Could not parse SP sheet: {e}")
    else:
        st.info(tr("no_data"))
    st.markdown("<br>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3b — PRICE ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
if "Price Analysis" in active_sections:
    season_label = " + ".join(sel_seasons) if len(sel_seasons) <= 4 else f"{len(sel_seasons)} seasons"
    sec(f"{tr('price_analysis')}  ·  {season_label}")

    if not df_sp.empty and C:
        try:
            df_p = df_sp.copy()

            # Price source: "Price FCA.1" in contract currency (EUR / PLN / USD)
            # Convert to EUR and USD using per-contract exchange rates:
            #   Ex rate EUR = PLN per 1 EUR  → EUR = PLN_price / ex_eur
            #   Ex rate USD = PLN per 1 USD  → USD = PLN_price / ex_usd
            #   EUR/USD     = cross rate      → USD = EUR_price * eurusd
            C_FCA1   = C["fca1"]
            C_EXUSD  = C["ex_usd"]
            C_EXEUR  = C["ex_eur"]
            C_EURUSD = C["eurusd"]

            for col in [C_FCA1, C_EXUSD, C_EXEUR, C_EURUSD]:
                df_p[col] = pd.to_numeric(df_p[col], errors="coerce")
            df_p[C["sold_mt"]] = pd.to_numeric(df_p[C["sold_mt"]], errors="coerce")

            def _to_eur(row):
                p = row[C_FCA1]
                if pd.isna(p): return None
                cur = str(row[C["currency"]]).strip().upper()
                if cur == "EUR": return p
                if cur == "PLN":
                    ex = row[C_EXEUR]
                    return p / ex if pd.notna(ex) and ex != 0 else None
                if cur == "USD":
                    eurusd = row[C_EURUSD]
                    return p / eurusd if pd.notna(eurusd) and eurusd != 0 else None
                return None

            def _to_usd(row):
                p = row[C_FCA1]
                if pd.isna(p): return None
                cur = str(row[C["currency"]]).strip().upper()
                if cur == "USD": return p
                if cur == "PLN":
                    ex = row[C_EXUSD]
                    return p / ex if pd.notna(ex) and ex != 0 else None
                if cur == "EUR":
                    eurusd = row[C_EURUSD]
                    return p * eurusd if pd.notna(eurusd) else None
                return None

            df_p["_price_eur"] = df_p.apply(_to_eur, axis=1)
            df_p["_price_usd"] = df_p.apply(_to_usd, axis=1)
            C_EUR, C_USD = "_price_eur", "_price_usd"

            df_p = df_p[df_p["_date"].notna() & df_p[C_FCA1].notna() & (df_p[C["sold_mt"]].fillna(0) > 0)]

            if not df_p.empty:
                # ── Avg price metrics ─────────────────────────────────────────
                # Simple (unweighted) average — one price per contract, matches chart visual
                avg_eur = df_p[C_EUR].dropna().mean()
                avg_usd = df_p[C_USD].dropna().mean()
                # Also compute quantity-weighted for reference
                def wavg(price_col):
                    w  = df_p[C["sold_mt"]].fillna(0)
                    p  = df_p[price_col].fillna(0)
                    tw = w.sum()
                    return (p * w).sum() / tw if tw > 0 else None
                wavg_eur = wavg(C_EUR)
                wavg_usd = wavg(C_USD)

                a1, a2, a3, a4 = st.columns(4)
                with a1: st.metric(tr("avg_price_eur"), f"€ {avg_eur:,.2f}" if pd.notna(avg_eur) else "—",
                                   help="Simple average: one price per contract")
                with a2: st.metric(tr("avg_price_usd"), f"$ {avg_usd:,.2f}" if pd.notna(avg_usd) else "—",
                                   help="Simple average: one price per contract")
                with a3: st.metric("Wtd Avg (EUR)", f"€ {wavg_eur:,.2f}" if wavg_eur else "—",
                                   help="Quantity-weighted average (larger contracts count more)")
                with a4: st.metric("Wtd Avg (USD)", f"$ {wavg_usd:,.2f}" if wavg_usd else "—",
                                   help="Quantity-weighted average (larger contracts count more)")

                st.markdown("<br>", unsafe_allow_html=True)

                # ── Chart controls ────────────────────────────────────────────
                ctrl1, ctrl2, ctrl3 = st.columns([2, 2, 2])
                with ctrl1:
                    view_mode = st.radio(
                        "Time range",
                        [tr("view_1y"), tr("view_5y"), tr("view_all")],
                        horizontal=True, key="price_view",
                    )
                with ctrl2:
                    chart_currency = st.radio(
                        tr("currency_toggle"), ["EUR", "USD"],
                        horizontal=True, key="price_currency",
                    )
                with ctrl3:
                    # Product breakdown toggle
                    show_by_product = st.checkbox("Split by product", value=False, key="price_split")

                # ── Build chart dataframe based on view mode ──────────────────
                price_col = C_EUR if chart_currency == "EUR" else C_USD
                currency_sym = "€" if chart_currency == "EUR" else "$"

                df_chart = df_p[["_date", price_col, C["sold_mt"], C["goods"]]].copy()
                df_chart.columns = ["date", "price", "qty", "product"]
                df_chart = df_chart.dropna(subset=["date","price"])
                df_chart["date"] = pd.to_datetime(df_chart["date"])
                df_chart = df_chart.sort_values("date")

                # Apply time range filter
                max_date = df_chart["date"].max()
                en_view = view_mode
                # normalise view mode to English key for comparison
                for lang_key, val in [("view_1y","1 Year"),("view_5y","5 Years"),("view_all","All")]:
                    if view_mode == tr(lang_key):
                        en_view = val
                        break

                if en_view == "1 Year":
                    cutoff = max_date - pd.DateOffset(years=1)
                    df_chart = df_chart[df_chart["date"] >= cutoff]
                    # Group by ~3-day bins
                    freq = "3D"
                elif en_view == "5 Years":
                    cutoff = max_date - pd.DateOffset(years=5)
                    df_chart = df_chart[df_chart["date"] >= cutoff]
                    # Group by quarter
                    freq = "QS"
                else:
                    # All data — group by month
                    freq = "MS"

                if df_chart.empty:
                    st.info("No price data for selected range.")
                else:
                    # Weighted average price per period (optionally per product)
                    df_chart["period"] = df_chart["date"].dt.to_period(
                        "3D" if freq == "3D" else ("Q" if freq == "QS" else "M")
                    ).dt.to_timestamp()

                    if show_by_product:
                        grouped = df_chart.groupby(["period","product"]).apply(
                            lambda g: (g["price"] * g["qty"].fillna(1)).sum() / g["qty"].fillna(1).sum()
                        ).reset_index(name="avg_price")
                        products_in_data = grouped["product"].unique()
                        fig = go.Figure()
                        for prod in products_in_data:
                            sub_g = grouped[grouped["product"] == prod].sort_values("period")
                            fig.add_trace(go.Scatter(
                                x=sub_g["period"], y=sub_g["avg_price"],
                                mode="lines+markers", name=str(prod),
                                line=dict(width=2),
                                hovertemplate=f"%{{x|%b %Y}}<br>{currency_sym}%{{y:,.2f}}<extra>%{{fullData.name}}</extra>",
                            ))
                    else:
                        grouped = df_chart.groupby("period").apply(
                            lambda g: (g["price"] * g["qty"].fillna(1)).sum() / g["qty"].fillna(1).sum()
                        ).reset_index(name="avg_price")
                        grouped = grouped.sort_values("period")
                        fig = go.Figure()
                        fig.add_trace(go.Scatter(
                            x=grouped["period"], y=grouped["avg_price"],
                            mode="lines+markers", name=f"Avg FCA ({chart_currency})",
                            line=dict(color="#3a7bd5", width=2.5),
                            fill="tozeroy",
                            fillcolor="rgba(58,123,213,0.08)",
                            hovertemplate=f"%{{x|%b %Y}}<br>{currency_sym}%{{y:,.2f}}<extra></extra>",
                        ))
                        # Add range slider band
                        if len(grouped) > 3:
                            rolling_avg = grouped["avg_price"].rolling(3, min_periods=1).mean()
                            fig.add_trace(go.Scatter(
                                x=grouped["period"], y=rolling_avg,
                                mode="lines", name="3-period MA",
                                line=dict(color="#e09030", width=1.5, dash="dot"),
                                hovertemplate=f"%{{x|%b %Y}}<br>MA: {currency_sym}%{{y:,.2f}}<extra></extra>",
                            ))

                    fig.update_layout(
                        title=dict(text=tr("price_chart_title"), font=dict(family="IBM Plex Mono", size=15, color="#1a5fa0")),
                        paper_bgcolor="#f5f4f0", plot_bgcolor="#eceae4",
                        font=dict(family="IBM Plex Sans", color="#1a1a1a", size=13),
                        xaxis=dict(
                            showgrid=True, gridcolor="#c8c4bc", gridwidth=1,
                            showline=True, linecolor="#b8b4ac",
                            tickfont=dict(size=12, color="#505050"),
                            title=None,
                            rangeslider=dict(visible=True, bgcolor="#eceae4", thickness=0.08),
                        ),
                        yaxis=dict(
                            showgrid=True, gridcolor="#c8c4bc", gridwidth=1,
                            showline=True, linecolor="#b8b4ac",
                            tickfont=dict(size=12, color="#505050"),
                            title=dict(text=f"Price ({chart_currency})", font=dict(size=12, color="#666666", family="IBM Plex Sans")),
                            tickprefix=currency_sym,
                            # Start y-axis near the data minimum, not zero
                            rangemode="normal",
                            range=[
                                grouped["avg_price"].min() * 0.90 if not show_by_product else None,
                                grouped["avg_price"].max() * 1.05 if not show_by_product else None,
                            ] if not show_by_product else None,
                        ),
                        legend=dict(font=dict(family="IBM Plex Mono", size=12, color="#2a2a2a"),
                                    bgcolor="#eceae4", bordercolor="#c8c4bc", borderwidth=1),
                        hovermode="x unified",
                        hoverlabel=dict(bgcolor="#ffffff", bordercolor="#c8c4bc",
                                        font=dict(family="IBM Plex Mono", size=12, color="#1a1a1a")),
                        margin=dict(t=60, b=60, l=60, r=20),
                        height=460,
                    )
                    # Buttons for 1Y / 5Y / All quick-select via plotly range
                    fig.update_xaxes(
                        rangeselector=dict(
                            buttons=[
                                dict(count=3, label="3M", step="month", stepmode="backward"),
                                dict(count=6, label="6M", step="month", stepmode="backward"),
                                dict(count=1, label="1Y", step="year", stepmode="backward"),
                                dict(count=5, label="5Y", step="year", stepmode="backward"),
                                dict(step="all", label="All"),
                            ],
                            bgcolor="#eceae4", activecolor="#1a5fa0",
                            font=dict(color="#1a1a1a", size=11, family="IBM Plex Mono"),
                            bordercolor="#c8c4bc",
                            x=0, y=1.08,
                        )
                    )
                    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": True})

                    # ── Excel download ────────────────────────────────────────
                    with st.expander("📥 Download price data as Excel"):
                        dl_col1, dl_col2 = st.columns([2, 3])
                        with dl_col1:
                            exp_period = st.radio(
                                "Aggregation period",
                                ["Daily", "Weekly", "Monthly"],
                                horizontal=True,
                                key="price_export_period",
                            )
                        with dl_col2:
                            try:
                                xl_price = _build_price_excel(
                                    df_chart, exp_period, chart_currency, show_by_product
                                )
                                fname = f"avg_price_{chart_currency}_{exp_period.lower()}.xlsx"
                                st.download_button(
                                    label=f"⬇ Download ({exp_period})",
                                    data=xl_price,
                                    file_name=fname,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="price_xl_dl",
                                )
                            except Exception as ex:
                                st.warning(f"Could not build price export: {ex}")

                    # ── Price distribution table by product ───────────────────
                    st.markdown("<br>", unsafe_allow_html=True)
                    sub("PRICE BREAKDOWN BY PRODUCT")
                    breakdown = df_p.groupby(C["goods"]).apply(lambda g: pd.Series({
                        "Contracts":     len(g),
                        "Sold (MT)":     g[C["sold_mt"]].sum(),
                        f"Avg EUR":       round((g[C_EUR].fillna(0) * g[C["sold_mt"]].fillna(0)).sum() / max(g[C["sold_mt"]].fillna(0).sum(), 0.001), 2),
                        f"Min EUR":       g[C_EUR].min(),
                        f"Max EUR":       g[C_EUR].max(),
                        f"Avg USD":       round((g[C_USD].fillna(0) * g[C["sold_mt"]].fillna(0)).sum() / max(g[C["sold_mt"]].fillna(0).sum(), 0.001), 2),
                        f"Min USD":       g[C_USD].min(),
                        f"Max USD":       g[C_USD].max(),
                    })).reset_index().rename(columns={C["goods"]: "Product"})
                    breakdown = breakdown[breakdown["Sold (MT)"] > 0].sort_values("Sold (MT)", ascending=False)
                    for col in ["Avg EUR","Min EUR","Max EUR","Avg USD","Min USD","Max USD"]:
                        if col in breakdown.columns:
                            breakdown[col] = breakdown[col].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "—")
                    breakdown["Sold (MT)"] = breakdown["Sold (MT)"].apply(lambda x: f"{x:,.2f}")
                    breakdown["Contracts"] = breakdown["Contracts"].astype(int)
                    st.dataframe(breakdown, use_container_width=True, hide_index=True)
            else:
                st.info(tr("no_data"))
        except Exception as e:
            st.warning(f"Could not build price analysis: {e}")
    else:
        st.info(tr("no_data"))
    st.markdown("<br>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — UNLOADING SCHEDULE
# ══════════════════════════════════════════════════════════════════════════════
if "Unloading Schedule" in active_sections:
    season_label = " + ".join(sel_seasons) if len(sel_seasons) <= 4 else f"{len(sel_seasons)} seasons"
    sec(f"{tr('unload_schedule')}  ·  {season_label}")
    st.markdown(f"<div style='font-size:1rem;color:#505050;margin-bottom:20px;font-weight:500;'>"
                f"{tr('schedule_desc')}</div>", unsafe_allow_html=True)
    try:
        if not df_sp.empty and C:
            # Dynamically detect date columns: any column whose name parses as a date (year >= 2020)
            # load_sp converts all column names to str, so datetime headers become e.g. "2025-08-01 00:00:00"
            date_cols = []  # list of (col_name, datetime_obj, "Mon YYYY" label)
            for col in df_sp.columns:
                try:
                    dt = pd.to_datetime(col, errors="coerce")
                    if pd.notna(dt) and dt.year >= 2020:
                        date_cols.append((col, dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0),
                                          dt.strftime("%b %Y")))
                except Exception:
                    pass
            # Sort by date and deduplicate labels (two dates in same month → merge)
            date_cols.sort(key=lambda x: x[1])
            seen_lbl: dict = {}
            merged_cols: list = []  # (display_label, datetime_obj, [col_names])
            for col_name, dt, lbl in date_cols:
                if lbl in seen_lbl:
                    seen_lbl[lbl][2].append(col_name)
                else:
                    entry = [lbl, dt, [col_name]]
                    seen_lbl[lbl] = entry
                    merged_cols.append(entry)

            if merged_cols:
                # Build one record per product: sum each date column group
                records = []
                for prod, grp in df_sp.groupby(C["goods"], sort=True):
                    row_data = {"Product": str(prod), "_dt_map": {}}
                    total = 0.0
                    has_data = False
                    for lbl, dt, col_names in merged_cols:
                        val = 0.0
                        for cn in col_names:
                            if cn in grp.columns:
                                v = pd.to_numeric(grp[cn], errors="coerce").sum()
                                if pd.notna(v):
                                    val += v
                        val = round(val, 1) if val != 0 else 0.0
                        row_data[lbl] = val if val else None
                        row_data["_dt_map"][lbl] = dt
                        if val:
                            total += val
                            has_data = True
                    row_data["_total"] = total
                    if has_data:
                        records.append(row_data)

                if records:
                    # Only show months that have at least one non-zero value across all products
                    active_labels = [lbl for lbl, _, _ in merged_cols
                                     if any(r.get(lbl) for r in records)]
                    rows_out = []
                    for r in records:
                        rd = {"Product": r["Product"]}
                        for lbl in active_labels:
                            val = r.get(lbl)
                            rd[lbl] = f"{val:,.2f}" if val else "—"
                        rd[tr("total")] = f"{r['_total']:,.2f}" if r["_total"] else "—"
                        rows_out.append(rd)
                    st.dataframe(pd.DataFrame(rows_out), use_container_width=True, hide_index=True)

                    sub(tr("upcoming"))
                    now = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                    upcoming = [lbl for lbl, dt, _ in merged_cols
                                if lbl in active_labels and dt >= now][:3]
                    if upcoming:
                        up_cols = st.columns(len(upcoming))
                        for col_ui, month in zip(up_cols, upcoming):
                            with col_ui:
                                total_mt, lines = 0.0, []
                                for r in records:
                                    val = r.get(month)
                                    if val:
                                        total_mt += val
                                        lines.append(f"{r['Product']}: {val:,.2f} MT")
                                detail = "".join(
                                    f"<div style='font-size:0.95rem;color:#505050;font-weight:500;"
                                    f"margin-bottom:5px;'>{ln}</div>" for ln in lines)
                                st.markdown(
                                    f"<div style='background:#eceae4;border:1px solid #c8c4bc;"
                                    f"border-radius:6px;padding:20px 22px;'>"
                                    f"<div style='font-family:IBM Plex Mono;font-size:1rem;color:#1a5fa0;"
                                    f"font-weight:700;letter-spacing:0.08em;margin-bottom:14px;"
                                    f"text-transform:uppercase;'>{month}</div>"
                                    f"<div style='font-family:IBM Plex Mono;font-size:1.7rem;color:#1a1a1a;"
                                    f"font-weight:700;margin-bottom:14px;'>{total_mt:,.2f} MT</div>"
                                    f"{detail}</div>", unsafe_allow_html=True)
                    else:
                        st.info(tr("no_upcoming"))
                else:
                    st.info(tr("no_data"))
            else:
                st.warning("No date columns found in SP sheet.")
        else:
            st.info(tr("no_data"))
    except Exception as e:
        st.warning(f"Could not load schedule: {e}")
    st.markdown("<br>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — UNLOADED (WAGI TOTAL)
# ══════════════════════════════════════════════════════════════════════════════
if "Unloaded (Wagi Total)" in active_sections:
    sec(tr("wagi_section"))
    st.markdown(f"<div style='font-size:1rem;color:#505050;margin-bottom:20px;font-weight:500;'>"
                f"{tr('wagi_desc')}</div>", unsafe_allow_html=True)

    if not df_w.empty and not df_sp.empty and C:
        try:
            # Build per-contract price lookup from SP using Price FCA.1 + exchange rates
            sp_contracts = df_sp_full[[
                C["contract"], C["fca1"], C["ex_eur"], C["ex_usd"], C["eurusd"],
                C["currency"], C["goods"], C["trader"], "_contract_key"
            ]].copy()
            for col in [C["fca1"], C["ex_eur"], C["ex_usd"], C["eurusd"]]:
                sp_contracts[col] = pd.to_numeric(sp_contracts[col], errors="coerce")

            # Compute EUR and USD price per contract row using the same conversion logic
            def _conv_eur(row):
                p, cur = row[C["fca1"]], str(row[C["currency"]]).strip().upper()
                if pd.isna(p): return None
                if cur == "EUR": return p
                if cur == "PLN":
                    ex = row[C["ex_eur"]]
                    return p / ex if pd.notna(ex) and ex != 0 else None
                if cur == "USD":
                    eurusd = row[C["eurusd"]]
                    return p / eurusd if pd.notna(eurusd) and eurusd != 0 else None
                return None

            def _conv_usd(row):
                p, cur = row[C["fca1"]], str(row[C["currency"]]).strip().upper()
                if pd.isna(p): return None
                if cur == "USD": return p
                if cur == "PLN":
                    ex = row[C["ex_usd"]]
                    return p / ex if pd.notna(ex) and ex != 0 else None
                if cur == "EUR":
                    eurusd = row[C["eurusd"]]
                    return p * eurusd if pd.notna(eurusd) else None
                return None

            sp_contracts["_price_eur_conv"] = sp_contracts.apply(_conv_eur, axis=1)
            sp_contracts["_price_usd_conv"] = sp_contracts.apply(_conv_usd, axis=1)

            # One row per contract — keep first non-null EUR price
            sp_contracts = sp_contracts.sort_values("_price_eur_conv", na_position="last")
            sp_contracts = sp_contracts.drop_duplicates(subset=["_contract_key"], keep="first")

            df_wm = df_w.merge(
                sp_contracts.rename(columns={
                    C["fca1"]:     "_price_fca",
                    C["currency"]: "_currency",
                    C["goods"]:    "_sp_goods",
                    C["trader"]:   "_sp_trader",
                }),
                on="_contract_key",
                how="left",
            )

            df_wm["_price_fca"]      = pd.to_numeric(df_wm["_price_fca"],      errors="coerce")
            df_wm["_price_eur_conv"] = pd.to_numeric(df_wm["_price_eur_conv"], errors="coerce")
            df_wm["_price_usd_conv"] = pd.to_numeric(df_wm["_price_usd_conv"], errors="coerce")

            # Value in EUR and USD using the contract's own exchange rate
            df_wm["_value_eur"] = df_wm["Qty_MT"] * df_wm["_price_eur_conv"]
            df_wm["_value_usd"] = df_wm["Qty_MT"] * df_wm["_price_usd_conv"]

            total_mt = df_wm["Qty_MT"].sum()
            n_rows   = len(df_wm)

            # Currency toggle for value display
            val_cur = st.radio(
                "Value display currency", ["EUR", "USD"],
                horizontal=True, key="wagi_val_currency",
            )
            total_val_eur = df_wm["_value_eur"].sum()
            total_val_usd = df_wm["_value_usd"].sum()

            # Top metrics
            t1, t2, t3 = st.columns(3)
            with t1:
                st.metric(tr("total_unloaded"), f"{total_mt:,.2f}")
                if st.button("🔍", key="drill_wagi"):
                    show_drill(df_wm[["Date_WZ","Product","Qty_MT","Season_col",
                                       "_contract_key","_price_eur_conv","_price_usd_conv",
                                       "_value_eur","_value_usd"]].rename(
                        columns={"Date_WZ":"Date","Qty_MT":"MT","Season_col":"Season",
                                 "_contract_key":"Contract","_price_eur_conv":"Price (EUR)",
                                 "_price_usd_conv":"Price (USD)",
                                 "_value_eur":"Value (EUR)","_value_usd":"Value (USD)"}
                    ).assign(Date=lambda d: pd.to_datetime(d["Date"],errors="coerce").dt.strftime("%d/%m/%Y")),
                    tr("wagi_section"))
                    st.rerun()
            with t2:
                st.metric(tr("shipments"), f"{n_rows:,}")
            with t3:
                if val_cur == "EUR":
                    st.metric(tr("unloaded_value"), f"€ {total_val_eur:,.2f}" if total_val_eur else "—",
                              help="Sum of: Qty (MT) × Price to EUR (rate on contract date)")
                else:
                    st.metric(tr("unloaded_value"), f"$ {total_val_usd:,.2f}" if total_val_usd else "—",
                              help="Sum of: Qty (MT) × Price to USD (rate on contract date)")

            st.markdown("<br>", unsafe_allow_html=True)

            # By product
            sub(tr("by_product"))
            by_prod = df_wm.groupby("Product").agg(
                MT=("Qty_MT", "sum"),
                Shipments=("Qty_MT", "count"),
            ).reset_index().sort_values("MT", ascending=False)
            by_prod_total = pd.DataFrame([{
                "Product": "TOTAL",
                "MT": by_prod["MT"].sum(),
                "Shipments": by_prod["Shipments"].sum(),
            }])
            by_prod = pd.concat([by_prod, by_prod_total], ignore_index=True)
            by_prod["MT"] = by_prod["MT"].apply(lambda x: f"{x:,.2f}")
            st.dataframe(by_prod, use_container_width=True, hide_index=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # By month
            sub(tr("by_month"))
            df_wm["YearMonth"] = df_wm["Date_WZ"].dt.to_period("M").astype(str)
            by_month = df_wm.groupby(["YearMonth","Product"])["Qty_MT"].sum().reset_index()
            pivot = by_month.pivot(index="YearMonth", columns="Product", values="Qty_MT").fillna(0)
            pivot[tr("total")] = pivot.sum(axis=1)
            pivot_disp = pivot.copy()
            for col in pivot_disp.columns:
                pivot_disp[col] = pivot_disp[col].apply(lambda x: f"{x:,.2f}" if x > 0 else "—")
            pivot_disp = pivot_disp.reset_index().rename(columns={"YearMonth": tr("month")})
            st.dataframe(pivot_disp, use_container_width=True, hide_index=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Recent
            sub(tr("recent_shipments"))
            recent = df_wm.sort_values("Date_WZ", ascending=False).head(50)
            show_r = recent[["Date_WZ","Product","Qty_MT","Season_col","_price_fca","_currency"]].copy()
            show_r.columns = ["Date", tr("product"), "MT", tr("season"), "Price FCA", "Currency"]
            show_r["Date"] = pd.to_datetime(show_r["Date"], errors="coerce").dt.strftime("%d %b %Y")
            show_r["MT"]   = show_r["MT"].apply(lambda x: f"{x:,.2f}")
            st.dataframe(show_r, use_container_width=True, hide_index=True, height=350)

        except Exception as e:
            st.warning(f"Could not load Wagi total: {e}")
    elif df_w.empty:
        st.info(tr("no_data"))
    else:
        st.info("SP sheet required for contract value lookup.")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — BROKER REPORT
# ══════════════════════════════════════════════════════════════════════════════
if "Broker Report" in active_sections:
    season_label = " + ".join(sel_seasons) if len(sel_seasons) <= 4 else f"{len(sel_seasons)} seasons"
    sec(f"{tr('broker_report')}  ·  {season_label}")

    if not df_sp.empty and C and C.get("broker") and C.get("commission"):
        try:
            df_br = df_sp.copy()
            # Fix rates stored with comma decimal separator (e.g. "1,5" → 1.5)
            df_br[C["commission"]] = pd.to_numeric(
                df_br[C["commission"]].astype(str).str.replace(",", ".", regex=False),
                errors="coerce"
            )
            df_br[C["sold_mt"]]   = pd.to_numeric(df_br[C["sold_mt"]],   errors="coerce").fillna(0)
            df_br[C["issued_mt"]] = pd.to_numeric(df_br[C["issued_mt"]], errors="coerce").fillna(0)
            # Blank broker = no broker used — exclude from report
            df_br = df_br[
                df_br[C["broker"]].astype(str).str.strip()
                .replace({"nan": "", "None": "", "NaN": ""}) != ""
            ]

            # ── Delivery date filter (overrides main date filter for this section) ──
            wagi_src = df_wagi_full if isinstance(df_wagi_full, pd.DataFrame) else None
            br_periods = []
            if wagi_src is not None and not wagi_src.empty:
                br_periods = sorted(
                    wagi_src["Date_WZ"].dropna().dt.to_period("M").unique(),
                    key=lambda p: p.ordinal
                )

            br_years  = ["All"] + sorted({p.year  for p in br_periods}, reverse=True)
            br_months = ["All"] + list(range(1, 13))
            MONTH_NAMES = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                           7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

            _brc1, _brc2 = st.columns(2)
            with _brc1:
                sel_br_year = st.selectbox(
                    "📦 Delivery year", br_years, key="br_delivery_year"
                )
            with _brc2:
                sel_br_month = st.selectbox(
                    "📦 Delivery month",
                    ["All"] + [MONTH_NAMES[m] for m in range(1, 13)],
                    key="br_delivery_month"
                )

            sel_br_month_num = next(
                (m for m, n in MONTH_NAMES.items() if n == sel_br_month), None
            )

            # Build a label and chosen period for downstream use
            if sel_br_year != "All" or sel_br_month != "All":
                _yr_str = str(sel_br_year) if sel_br_year != "All" else ""
                _mo_str = sel_br_month   if sel_br_month != "All" else ""
                sel_br_period = f"{_mo_str} {_yr_str}".strip()
            else:
                sel_br_period = "All"

            chosen_period = None
            if sel_br_period != "All" and wagi_src is not None:
                # Find matching period(s) — filter wagi by year and/or month
                def _period_match(p):
                    if sel_br_year != "All" and p.year != int(sel_br_year):
                        return False
                    if sel_br_month_num is not None and p.month != sel_br_month_num:
                        return False
                    return True
                matched = [p for p in br_periods if _period_match(p)]
                # chosen_period used only for single-month case; keep first match for label
                chosen_period = matched[0] if len(matched) == 1 else ("multi" if matched else None)
                if matched:
                    wagi_month = wagi_src[
                        wagi_src["Date_WZ"].dt.to_period("M").isin(matched)
                    ]
                    keys_in_month = set(wagi_month["_contract_key"].dropna().astype(str))
                    df_br = df_br[df_br["_contract_key"].isin(keys_in_month)]
                    # Delivery MT = actual MT delivered in the selected month per contract
                    mt_in_month = (
                        wagi_month.groupby("_contract_key")["Qty_MT"].sum()
                        .rename("_delivery_mt").reset_index()
                    )
                    mt_in_month["_contract_key"] = mt_in_month["_contract_key"].astype(str)
                    df_br = df_br.merge(mt_in_month, on="_contract_key", how="left")
                    df_br["_delivery_mt"] = df_br["_delivery_mt"].fillna(0)
            if "_delivery_mt" not in df_br.columns:
                # "All" — use total issued MT as the commission base
                df_br["_delivery_mt"] = df_br[C["issued_mt"]]

            # Amount due = delivered weight in period × commission rate
            df_br["_commission_eur"] = df_br[C["commission"]] * df_br["_delivery_mt"]

            # Deduplicate contracts that appear multiple times across seasons
            # (same contract split into S24+S25 rows, etc.)
            # SP quantities (sold/issued/left) are per-season → sum them
            # Wagi delivery figures come from one physical delivery and are already
            # identical on every duplicated row → take first, NOT sum
            if C.get("contract") and C["contract"] in df_br.columns:
                _sum_cols   = [c for c in [C["sold_mt"], C["issued_mt"], C["left_mt"]]
                               if c in df_br.columns]
                _first_cols = [c for c in [C["date"], C["buyer"], C["goods"], C["broker"],
                                           C["commission"], C["currency"], "_buyer_abbrev",
                                           "_contract_key", "_delivery_mt", "_commission_eur"]
                               if c in df_br.columns]
                _agg = {c: "sum"   for c in _sum_cols}
                _agg.update({c: "first" for c in _first_cols})
                df_br = df_br.groupby(C["contract"], as_index=False).agg(_agg)

            delivered_label = f"Delivered (MT){' · ' + sel_br_period if sel_br_period != 'All' else ''}"

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Summary by broker ──────────────────────────────────────────
            sub("BY BROKER")
            broker_col = df_br[C["broker"]].fillna("Unknown").astype(str).str.strip()
            broker_summary = df_br.assign(_broker=broker_col).groupby("_broker").agg(
                Contracts      =(C["contract"],   "count"),
                Sold_MT        =(C["sold_mt"],    "sum"),
                Delivered_MT   =("_delivery_mt",  "sum"),
                Avg_Rate       =(C["commission"], "mean"),
                Commission_EUR =("_commission_eur","sum"),
            ).reset_index().rename(columns={"_broker": "Broker"})
            broker_summary = broker_summary[broker_summary["Sold_MT"] > 0].sort_values(
                "Commission_EUR", ascending=False
            )
            # Total row
            br_total = {"Broker": "TOTAL",
                        "Contracts": broker_summary["Contracts"].sum(),
                        "Sold_MT": broker_summary["Sold_MT"].sum(),
                        "Delivered_MT": broker_summary["Delivered_MT"].sum(),
                        "Avg_Rate": None,
                        "Commission_EUR": broker_summary["Commission_EUR"].sum()}
            broker_summary_disp = pd.concat(
                [broker_summary, pd.DataFrame([br_total])], ignore_index=True
            ).rename(columns={
                "Contracts":     "Contracts",
                "Sold_MT":       "Sold (MT)",
                "Delivered_MT":  delivered_label,
                "Avg_Rate":      tr("commission_rate"),
                "Commission_EUR":tr("commission_total"),
            })
            st.dataframe(
                broker_summary_disp.style.format({
                    "Sold (MT)":            "{:,.2f}",
                    delivered_label:        "{:,.2f}",
                    tr("commission_rate"):  lambda x: f"{x:.4f}" if pd.notna(x) else "—",
                    tr("commission_total"): lambda x: f"€ {x:,.2f}" if pd.notna(x) else "—",
                }),
                use_container_width=True, hide_index=True,
            )

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Per-contract detail table ──────────────────────────────────
            sub("CONTRACT DETAILS")
            detail = df_br[[
                C["date"], C["contract"], C["buyer"],
                C["goods"], C["sold_mt"], C["broker"], C["commission"], C["currency"],
            ]].copy()
            detail.columns = [
                "Date", "Contract", "Client", "Goods",
                "Sold (MT)", "Broker", tr("commission_rate"), "Currency",
            ]
            detail[delivered_label]  = df_br["_delivery_mt"].values
            detail[tr("amount_due")] = df_br["_commission_eur"].values
            detail["Date"] = pd.to_datetime(detail["Date"], errors="coerce").dt.strftime("%d/%m/%Y")
            detail = detail[detail["Sold (MT)"] > 0].sort_values(["Broker", "Date"])
            # Reorder so delivered and amount due sit after sold
            cols = ["Date", "Contract", "Client", "Goods", "Sold (MT)",
                    delivered_label, tr("commission_rate"), tr("amount_due"), "Broker", "Currency"]
            detail = detail[[c for c in cols if c in detail.columns]]
            st.dataframe(
                detail.style.format({
                    "Sold (MT)":          "{:,.2f}",
                    "Delivered (MT)":     "{:,.2f}",
                    "Left (MT)":          "{:,.2f}",
                    tr("commission_rate"): "{:.4f}",
                    tr("amount_due"):      "€ {:,.2f}",
                }, na_rep="—"),
                use_container_width=True, hide_index=True,
            )

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Excel export ───────────────────────────────────────────────
            sub("EXPORT")
            fname = f"Broker_Report_{datetime.now().strftime('%B_%Y')}.xlsx"
            wagi_src = df_wagi_full if isinstance(df_wagi_full, pd.DataFrame) else None
            try:
                xl_bytes = _build_broker_excel(df_br, wagi_src, C)
                st.download_button(
                    label    = tr("dl_broker"),
                    data     = xl_bytes,
                    file_name= fname,
                    mime     = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key      = "broker_dl_btn",
                )
            except Exception as xl_err:
                st.warning(f"Could not generate Excel: {xl_err}")

        except Exception as e:
            st.warning(f"Could not build broker report: {e}")
    elif not C.get("broker"):
        st.info("No 'Broker' column found in the SP sheet.")
    elif not C.get("commission"):
        st.info("No 'Commision' / 'Commission' column found in the SP sheet.")
    else:
        st.info(tr("no_data"))
    st.markdown("<br>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — BUYER SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
if "Buyer Summary" in active_sections:
    season_label = " + ".join(sel_seasons) if len(sel_seasons) <= 4 else f"{len(sel_seasons)} seasons"
    sec(f"{tr('buyer_summary')}  ·  {season_label}")

    if not df_sp.empty and C and C.get("buyer"):
        try:
            df_bs = df_sp.copy()

            # Per-row converted values (same logic as Contract Summary)
            def _bs_to_eur(row):
                p   = pd.to_numeric(row.get(C["fca1"]),    errors="coerce")
                q   = pd.to_numeric(row.get(C["sold_mt"]), errors="coerce")
                cur = str(row.get(C["currency"], "")).strip().upper()
                if pd.isna(p) or pd.isna(q) or q == 0: return 0.0
                if cur == "EUR": return p * q
                ex_eur = pd.to_numeric(row.get(C["ex_eur"]), errors="coerce")
                eurusd = pd.to_numeric(row.get(C["eurusd"]), errors="coerce")
                if cur == "PLN" and pd.notna(ex_eur) and ex_eur != 0: return (p / ex_eur) * q
                if cur == "USD" and pd.notna(eurusd)  and eurusd != 0: return (p / eurusd)  * q
                return 0.0

            def _bs_to_usd(row):
                p   = pd.to_numeric(row.get(C["fca1"]),    errors="coerce")
                q   = pd.to_numeric(row.get(C["sold_mt"]), errors="coerce")
                cur = str(row.get(C["currency"], "")).strip().upper()
                if pd.isna(p) or pd.isna(q) or q == 0: return 0.0
                if cur == "USD": return p * q
                ex_usd = pd.to_numeric(row.get(C["ex_usd"]), errors="coerce")
                eurusd = pd.to_numeric(row.get(C["eurusd"]), errors="coerce")
                if cur == "PLN" and pd.notna(ex_usd) and ex_usd != 0: return (p / ex_usd) * q
                if cur == "EUR" and pd.notna(eurusd):                   return p * eurusd  * q
                return 0.0

            def _bs_to_pln(row):
                p   = pd.to_numeric(row.get(C["fca1"]),    errors="coerce")
                q   = pd.to_numeric(row.get(C["sold_mt"]), errors="coerce")
                cur = str(row.get(C["currency"], "")).strip().upper()
                if pd.isna(p) or pd.isna(q) or q == 0: return 0.0
                if cur == "PLN": return p * q
                ex_eur = pd.to_numeric(row.get(C["ex_eur"]), errors="coerce")
                ex_usd = pd.to_numeric(row.get(C["ex_usd"]), errors="coerce")
                if cur == "EUR" and pd.notna(ex_eur): return p * ex_eur * q
                if cur == "USD" and pd.notna(ex_usd): return p * ex_usd * q
                return 0.0

            df_bs["_r_eur"] = df_bs.apply(_bs_to_eur, axis=1)
            df_bs["_r_usd"] = df_bs.apply(_bs_to_usd, axis=1)
            df_bs["_r_pln"] = df_bs.apply(_bs_to_pln, axis=1)
            df_bs[C["sold_mt"]]   = pd.to_numeric(df_bs[C["sold_mt"]],   errors="coerce").fillna(0)
            df_bs[C["issued_mt"]] = pd.to_numeric(df_bs[C["issued_mt"]], errors="coerce").fillna(0)
            df_bs[C["left_mt"]]   = pd.to_numeric(df_bs[C["left_mt"]],   errors="coerce").fillna(0)

            buyer_grp = df_bs.groupby("_buyer_abbrev", as_index=False).apply(
                lambda g: pd.Series({
                    "_full_name":  g[C["buyer"]].dropna().iloc[0] if not g[C["buyer"]].dropna().empty else "",
                    "Contracts":   len(g),
                    "Sold (MT)":   g[C["sold_mt"]].sum(),
                    "Issued (MT)": g[C["issued_mt"]].sum(),
                    "Left (MT)":   g[C["left_mt"]].sum(),
                    "Value → EUR": g["_r_eur"].sum(),
                    "Value → USD": g["_r_usd"].sum(),
                    "Value → PLN": g["_r_pln"].sum(),
                })
            ).rename(columns={"_buyer_abbrev": "Buyer"})

            buyer_grp = (buyer_grp[buyer_grp["Sold (MT)"] > 0]
                         .sort_values("Sold (MT)", ascending=False)
                         .reset_index(drop=True))

            # Avg price per currency = total value / total MT (quantity-weighted)
            buyer_grp["Avg EUR"] = (buyer_grp["Value → EUR"] / buyer_grp["Sold (MT)"]).where(buyer_grp["Sold (MT)"] > 0)
            buyer_grp["Avg USD"] = (buyer_grp["Value → USD"] / buyer_grp["Sold (MT)"]).where(buyer_grp["Sold (MT)"] > 0)
            buyer_grp["Avg PLN"] = (buyer_grp["Value → PLN"] / buyer_grp["Sold (MT)"]).where(buyer_grp["Sold (MT)"] > 0)

            # ── Currency selector ──────────────────────────────────────────
            bs_cur = st.radio(
                "Avg Price FCA currency", ["EUR", "USD", "PLN"],
                horizontal=True, key="bs_avg_cur"
            )
            avg_col    = {"EUR": "Avg EUR", "USD": "Avg USD", "PLN": "Avg PLN"}[bs_cur]
            avg_sym    = {"EUR": "€", "USD": "$", "PLN": "zł"}[bs_cur]
            avg_label  = f"Avg Price FCA ({bs_cur})"

            # ── Display table with total row ───────────────────────────────
            num_cols_bs = ["Contracts","Sold (MT)","Issued (MT)","Left (MT)",
                           "Value → EUR","Value → USD","Value → PLN"]
            total_bs = {"Buyer": "TOTAL", "_full_name": ""}
            for col in num_cols_bs:
                total_bs[col] = buyer_grp[col].sum()
            # Total avg price = total value / total MT
            _tot_mt = buyer_grp["Sold (MT)"].sum()
            total_bs["Avg EUR"] = buyer_grp["Value → EUR"].sum() / _tot_mt if _tot_mt else None
            total_bs["Avg USD"] = buyer_grp["Value → USD"].sum() / _tot_mt if _tot_mt else None
            total_bs["Avg PLN"] = buyer_grp["Value → PLN"].sum() / _tot_mt if _tot_mt else None

            disp_bs = pd.concat([buyer_grp, pd.DataFrame([total_bs])], ignore_index=True)

            # Build display columns — swap in the selected avg column
            display_cols = ["Buyer","_full_name","Contracts","Sold (MT)","Issued (MT)",
                            "Left (MT)","Value → EUR","Value → USD","Value → PLN", avg_col]
            disp_show = disp_bs[display_cols].rename(columns={
                "_full_name": "Full Name",
                avg_col: avg_label,
            })
            st.dataframe(
                disp_show.style.format({
                    "Sold (MT)":   "{:,.2f}", "Issued (MT)": "{:,.2f}",
                    "Left (MT)":   "{:,.2f}",
                    "Value → EUR": "€ {:,.2f}", "Value → USD": "$ {:,.2f}",
                    "Value → PLN": "zł {:,.2f}",
                    avg_label: lambda x: f"{avg_sym} {x:,.2f}" if pd.notna(x) and x else "—",
                }, na_rep="—"),
                use_container_width=True, hide_index=True,
            )

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Excel export (includes all three avg price columns) ────────
            sub("EXPORT")
            fname_bs = f"Buyer_Summary_{datetime.now().strftime('%B_%Y')}.xlsx"
            if st.button(tr("export_buyer"), key="buyer_export_btn"):
                export_df = disp_bs.rename(columns={
                    "_full_name": "Full Name",
                    "Avg EUR": "Avg Price EUR",
                    "Avg USD": "Avg Price USD",
                    "Avg PLN": "Avg Price PLN",
                })[["Buyer","Full Name","Contracts","Sold (MT)","Issued (MT)",
                    "Left (MT)","Value → EUR","Value → USD","Value → PLN",
                    "Avg Price EUR","Avg Price USD","Avg Price PLN"]]
                xl_bytes = _build_buyer_excel(export_df)
                st.download_button(
                    label     = tr("dl_buyer"),
                    data      = xl_bytes,
                    file_name = fname_bs,
                    mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key       = "buyer_dl_btn",
                )

        except Exception as e:
            st.warning(f"Could not build buyer summary: {e}")
    else:
        st.info(tr("no_data"))
    st.markdown("<br>", unsafe_allow_html=True)
